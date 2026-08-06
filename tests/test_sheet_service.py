import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from processing_cancellation import ProcessingCancellation  # noqa: E402
from services.sheet_service import merge_workbooks  # noqa: E402


def _books(root):
    template = root / "template.xlsx"
    source = root / "source.xlsx"
    template_book = Workbook()
    template_sheet = template_book.active
    template_sheet.title = "Data"
    template_sheet.append(["name", "value"])
    template_sheet.append(["existing", 1])
    template_book.save(template)
    source_book = Workbook()
    source_sheet = source_book.active
    source_sheet.title = "Data"
    source_sheet.append(["name", "value"])
    source_sheet.append(["new", 2])
    source_book.save(source)
    return template, source


class SheetServiceTests(unittest.TestCase):
    def test_appends_source_rows_without_overwriting_template(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            template, source = _books(root)
            output = root / "outputs"
            settings = {
                "Data": {
                    "S": 1,
                    "E": 1,
                    "mode": 1,
                    "key_col": "",
                    "protect": True,
                }
            }
            result = merge_workbooks(
                str(template), [str(source)], None, str(output), settings=settings
            )
            self.assertEqual(result.state.value, "succeeded")
            self.assertEqual(len(result.output_files), 1)
            result_book = load_workbook(result.output_files[0], data_only=True)
            self.assertEqual(result_book["Data"].cell(3, 1).value, "new")
            result_book.close()
            original_book = load_workbook(template, data_only=True)
            self.assertEqual(original_book["Data"].max_row, 2)
            original_book.close()

    def test_cancelled_job_does_not_save_result(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            template, source = _books(root)
            output = root / "outputs"
            cancellation = ProcessingCancellation()

            def report(event):
                if event.current_sheet == "Data":
                    cancellation.request_cancel_all()

            result = merge_workbooks(
                str(template),
                [str(source)],
                None,
                str(output),
                report,
                settings={
                    "Data": {
                        "S": 1,
                        "E": 1,
                        "mode": 1,
                        "key_col": "",
                        "protect": True,
                    }
                },
                cancellation=cancellation,
            )
            self.assertEqual(result.state.value, "cancelled")
            self.assertEqual(list(output.glob("*.xlsx")), [])

    def test_progress_reports_current_sheet(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            template, source = _books(root)
            events = []
            merge_workbooks(
                str(template),
                [str(source)],
                None,
                str(root / "outputs"),
                events.append,
                settings={
                    "Data": {
                        "S": 1,
                        "E": 1,
                        "mode": 1,
                        "key_col": "",
                        "protect": True,
                    }
                },
            )
            self.assertIn("Data", [event.current_sheet for event in events])

    def test_progress_uses_file_samples_without_counting_sheet_events(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            template, source = _books(root)
            source2 = root / "source2.xlsx"
            source2.write_bytes(source.read_bytes())
            events = []
            result = merge_workbooks(
                str(template),
                [str(source), str(source2)],
                None,
                str(root / "outputs"),
                events.append,
                settings={
                    "Data": {"S": 1, "E": 1, "mode": 1, "key_col": "", "protect": True}
                },
            )
            sheet_events = [event for event in events if event.current_sheet == "Data"]
            self.assertEqual([event.completed for event in sheet_events], [0, 1])
            first_complete = next(event for event in events if event.completed == 1 and not event.current_sheet)
            self.assertEqual(first_complete.estimate_status, "available")
            self.assertIn("xlsx_merge", result.details["timing"]["sample_counts"])


if __name__ == "__main__":
    unittest.main()
