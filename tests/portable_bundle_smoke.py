import json
import subprocess
import time
import zipfile
from pathlib import Path
from openpyxl import Workbook, load_workbook
import win32gui

source_root = Path(__file__).resolve().parents[1]
project_root = source_root.parent
archive = project_root / "release" / "DnS_Auto_Portable.zip"
destination = project_root / f"통합_포터블_이동테스트_{time.time_ns()}"
destination.mkdir()
with zipfile.ZipFile(archive) as package:
    package.extractall(destination)
bundle = destination / "DnS Auto"
inputs = bundle / "inputs"
profile_path = bundle / "profiles" / "sheet" / "profile.json"
mcp_exe = bundle / "DnS Auto MCP.exe"
gui_exe = bundle / "DnS Auto.exe"
for required in (mcp_exe, gui_exe, bundle / "USER_GUIDE.html", bundle / "GUI_GUIDE.html", bundle / "MCP_GUIDE.html", bundle / "QUICK_START.txt"):
    assert required.is_file(), required

# 이동된 GUI EXE가 실제 메인 창을 표시하는지 확인합니다.
existing = set()
def remember(handle, _extra):
    if "DnS Auto" in win32gui.GetWindowText(handle):
        existing.add(handle)
win32gui.EnumWindows(remember, None)
gui = subprocess.Popen([str(gui_exe)], cwd=Path.home())
window_title = ""
try:
    deadline = time.monotonic() + 30
    while time.monotonic() < deadline:
        found = []
        def collect(handle, _extra):
            title = win32gui.GetWindowText(handle)
            if handle not in existing and "DnS Auto" in title:
                found.append(title)
        win32gui.EnumWindows(collect, None)
        if found:
            window_title = found[0]
            break
        if gui.poll() is not None:
            break
        time.sleep(0.1)
    assert window_title, f"GUI window not found; exit={gui.poll()}"
    assert "v1.0.0" in window_title, window_title
finally:
    gui.terminate()
    try:
        gui.wait(timeout=10)
    except subprocess.TimeoutExpired:
        gui.kill()
        gui.wait(timeout=5)

def workbook(path, rows):
    book = Workbook()
    sheet = book.active
    sheet.title = "Data"
    for row in rows:
        sheet.append(row)
    book.save(path)

workbook(inputs / "template.xlsx", [["name", "value"], ["existing", 1]])
workbook(inputs / "source.xlsx", [["name", "value"], ["new", 2]])
profile = {"profile_type": "sheet_config", "sheet_configs": [{"sheet_name": "Data", "header_start": 1, "header_end": 1, "mode": 1, "key_col": "", "protect": True}]}
profile_path.write_text(json.dumps(profile, ensure_ascii=False), encoding="utf-8")

process = subprocess.Popen([str(mcp_exe)], cwd=Path.home(), stdin=subprocess.PIPE, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, encoding="utf-8")
def request(identifier, method, params=None):
    payload = {"jsonrpc": "2.0", "id": identifier, "method": method}
    if params is not None:
        payload["params"] = params
    process.stdin.write(json.dumps(payload, ensure_ascii=False) + "\n")
    process.stdin.flush()
    line = process.stdout.readline()
    if not line:
        raise RuntimeError(process.stderr.read())
    return json.loads(line)

try:
    initialized = request(1, "initialize", {"protocolVersion": "2024-11-05"})
    assert initialized["result"]["serverInfo"]["name"] == "dns-auto-mcp"
    assert initialized["result"]["serverInfo"]["version"] == "1.0.0"
    tools = request(2, "tools/list")["result"]["tools"]
    assert any(tool["name"] == "discover_merge_plan" for tool in tools)
    assert any(tool["name"] == "cancel_job" for tool in tools)
    discovered = request(3, "tools/call", {"name": "discover_merge_plan", "arguments": {"profile_name": "profile"}})
    plan = json.loads(discovered["result"]["content"][0]["text"])
    assert plan["status"] == "ready" and plan["operation"] == "excel", plan
    assert plan["selected_template"].endswith("template.xlsx"), plan
    assert any(path.endswith("source.xlsx") for path in plan["source_candidates"]), plan
    for path in (inputs / "source.xlsx", profile_path):
        inspected = request(2, "tools/call", {"name": "inspect_files", "arguments": {"paths": [str(path)]}})
        assert json.loads(inspected["result"]["content"][0]["text"])["files"][0]["ok"], path
    started = request(3, "tools/call", {"name": "start_sheet_merge", "arguments": {"template_path": str(inputs / "template.xlsx"), "source_paths": [str(inputs / "source.xlsx")], "profile_path": str(profile_path)}})
    job_id = json.loads(started["result"]["content"][0]["text"])["job_id"]
    result = None
    for attempt in range(100):
        status = request(4 + attempt, "tools/call", {"name": "get_job_result", "arguments": {"job_id": job_id}})
        result = json.loads(status["result"]["content"][0]["text"])
        if result.get("result") is not None:
            break
        time.sleep(0.05)
    assert result["state"] == "succeeded", result
    output = Path(result["result"]["output_files"][0])
    assert output.parent == (bundle / "outputs").resolve(), output
    result_book = load_workbook(output, data_only=True)
    assert result_book["Data"].cell(3, 1).value == "new"
    print(json.dumps({"bundle": str(bundle), "gui_window": window_title, "shared_profile": str(profile_path), "mcp_output": str(output), "state": result["state"]}, ensure_ascii=False))
finally:
    process.terminate()
    process.wait(timeout=10)
