"""공식 DOCX/XLS 자료로 COM 종료 패치와 잔존 프로세스를 비식별 검증한다."""

from __future__ import annotations

import json
import sys
import time
from datetime import datetime
from pathlib import Path

import fitz
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(ROOT))
from com_process_ownership import capture_processes  # noqa: E402
from services.conversion_service import convert_docx_to_pdf, convert_xls_to_xlsx  # noqa: E402


def process_counts():
    return {
        "word": len(capture_processes(["WINWORD.EXE"])),
        "excel": len(capture_processes(["EXCEL.EXE"])),
    }


def main():
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    root = ROOT / "tests" / "results" / "com_cleanup" / f"actual-{stamp}"
    root.mkdir(parents=True)
    docx = sorted((ROOT / "tests" / "test_files" / "PDF 취합 테스트" / "docx 변환하기").glob("*.docx"))
    xls = sorted((ROOT / "tests" / "test_files" / "엑셀 취합 테스트" / "6. xls변환하기").glob("*.xls"))
    before = process_counts()
    started = time.perf_counter()
    docx_result = convert_docx_to_pdf([str(path) for path in docx], str(root / "docx"))
    docx_elapsed = time.perf_counter() - started
    started = time.perf_counter()
    xls_result = convert_xls_to_xlsx([str(path) for path in xls], str(root / "xls"))
    xls_elapsed = time.perf_counter() - started
    time.sleep(1)
    after = process_counts()
    pages = 0
    for output in docx_result.output_files:
        with fitz.open(output) as document:
            pages += document.page_count
    readable_workbooks = 0
    for output in xls_result.output_files:
        book = load_workbook(output, read_only=True, data_only=False)
        readable_workbooks += int(bool(book.sheetnames))
        book.close()
    payload = {
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "process_counts_before": before,
        "process_counts_after": after,
        "docx": {
            "elapsed_seconds": round(docx_elapsed, 3),
            "state": docx_result.state.value,
            "input_count": len(docx),
            "output_count": len(docx_result.output_files),
            "failed_count": len(docx_result.failed_files),
            "total_pages": pages,
            "com_cleanup": docx_result.details.get("com_cleanup"),
        },
        "xls": {
            "elapsed_seconds": round(xls_elapsed, 3),
            "state": xls_result.state.value,
            "input_count": len(xls),
            "output_count": len(xls_result.output_files),
            "failed_count": len(xls_result.failed_files),
            "readable_workbook_count": readable_workbooks,
            "com_cleanup": xls_result.details.get("com_cleanup"),
        },
    }
    path = root / "summary.json"
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"result_root": str(root), **payload}, ensure_ascii=False))


if __name__ == "__main__":
    main()
