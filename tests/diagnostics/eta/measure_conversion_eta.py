"""ETA 패치 전후 DOCX/XLS 변환 실측. 입력 원본은 읽기만 한다."""

from __future__ import annotations

import importlib.util
import json
import sys
import time
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(ROOT))

from services import conversion_service as patched  # noqa: E402


def load_baseline():
    path = ROOT / "백업" / "20260805-001" / "services" / "conversion_service.py"
    spec = importlib.util.spec_from_file_location("eta_baseline_conversion_service", path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def workbook_values(path):
    book = load_workbook(path, data_only=False, read_only=True)
    try:
        return {
            sheet.title: tuple(tuple(cell.value for cell in row) for row in sheet.iter_rows())
            for sheet in book.worksheets
        }
    finally:
        book.close()


def compare_xlsx(left_dir, right_dir):
    left = {path.name: path for path in left_dir.glob("*.xlsx")}
    right = {path.name: path for path in right_dir.glob("*.xlsx")}
    if left.keys() != right.keys():
        return False, sorted(set(left) ^ set(right))
    differences = []
    for name in sorted(left):
        if workbook_values(left[name]) != workbook_values(right[name]):
            differences.append(name)
    return not differences, differences


def run_case(module, function_name, sources, output_dir):
    events = []
    started = time.perf_counter()
    result = getattr(module, function_name)(
        [str(path) for path in sources], str(output_dir), events.append
    )
    elapsed = time.perf_counter() - started
    outputs = [Path(path) for path in result.output_files]
    eta_events = [event for event in events if event.estimated_remaining_seconds is not None]
    return {
        "elapsed_seconds": round(elapsed, 3),
        "state": result.state.value,
        "input_count": len(sources),
        "output_count": len(outputs),
        "failed_count": len(result.failed_files),
        "nonempty_output_count": sum(path.is_file() and path.stat().st_size > 0 for path in outputs),
        "event_count": len(events),
        "eta_available_event_count": len(eta_events),
        "first_eta_completed": eta_events[0].completed if eta_events else None,
        "timing": result.details.get("timing"),
    }


def main():
    baseline = load_baseline()
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    result_root = ROOT / "tests" / "results" / "eta" / f"measurement-{stamp}"
    result_root.mkdir(parents=True, exist_ok=False)
    xls_sources = sorted(
        (ROOT / "tests" / "test_files" / "엑셀 취합 테스트" / "6. xls변환하기").glob("*.xls")
    )
    docx_sources = sorted(
        (ROOT / "tests" / "test_files" / "PDF 취합 테스트" / "docx 변환하기").glob("*.docx")
    )
    report = {"created_at": datetime.now().isoformat(timespec="seconds"), "runs": {}}
    for kind, function_name, sources in (
        ("xls", "convert_xls_to_xlsx", xls_sources),
        ("docx", "convert_docx_to_pdf", docx_sources),
    ):
        for label, module in (("baseline", baseline), ("patched", patched)):
            output_dir = result_root / label / kind
            output_dir.mkdir(parents=True)
            try:
                report["runs"][f"{label}_{kind}"] = run_case(
                    module, function_name, sources, output_dir
                )
            except Exception as error:
                report["runs"][f"{label}_{kind}"] = {
                    "error_type": type(error).__name__,
                    "error": str(error),
                }
            (result_root / "measurement.partial.json").write_text(
                json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8"
            )
    equal, differences = compare_xlsx(
        result_root / "baseline" / "xls", result_root / "patched" / "xls"
    )
    report["validation"] = {
        "xls_cell_content_equal": equal,
        "xls_difference_count": len(differences),
        "docx_output_name_sets_equal": {
            path.name for path in (result_root / "baseline" / "docx").glob("*.pdf")
        }
        == {
            path.name for path in (result_root / "patched" / "docx").glob("*.pdf")
        },
    }
    report_path = result_root / "measurement.json"
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"result_root": str(result_root), **report}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
