"""PDF·Excel 설정 창에서 공통으로 사용하는 시트 미리보기 구성요소."""
# Copyright (C) 2026 두부코드(DOOBOO_CODE)
# SPDX-License-Identifier: AGPL-3.0-only

import tkinter as tk
from tkinter import ttk

from openpyxl.utils import get_column_letter


HEADER_PREVIEW_MARGIN = 10
HEADER_PREVIEW_DEBOUNCE_MS = 250
EXCEL_MAX_ROW = 1_048_576
MAX_PREVIEW_COLUMNS = 100


def preview_bounds(start_row, end_row=None):
    """S-10부터 E+10(또는 S+10)까지 Excel 행 경계 안에서 계산합니다."""
    low = max(1, start_row - HEADER_PREVIEW_MARGIN)
    high_base = end_row if end_row is not None else start_row
    return low, min(EXCEL_MAX_ROW, high_base + HEADER_PREVIEW_MARGIN)


def build_merged_value_lookup(
    worksheet,
    minimum_row=None,
    maximum_row=None,
    maximum_column=None,
):
    """병합 영역 좌표를 좌상단 값에 연결하되 필요한 표시 범위만 구성합니다."""
    lookup = {}
    for merged_range in worksheet.merged_cells.ranges:
        if minimum_row is not None and merged_range.max_row < minimum_row:
            continue
        if maximum_row is not None and merged_range.min_row > maximum_row:
            continue
        if maximum_column is not None and merged_range.min_col > maximum_column:
            continue
        value = worksheet.cell(
            row=merged_range.min_row,
            column=merged_range.min_col,
        ).value
        row_start = max(merged_range.min_row, minimum_row or merged_range.min_row)
        row_end = min(merged_range.max_row, maximum_row or merged_range.max_row)
        column_end = min(
            merged_range.max_col,
            maximum_column or merged_range.max_col,
        )
        for row in range(row_start, row_end + 1):
            for column in range(merged_range.min_col, column_end + 1):
                lookup[(row, column)] = value
    return lookup


class SheetPreviewPanel:
    """워크시트 일부를 색상 태그와 스크롤이 있는 Treeview로 표시합니다."""

    def __init__(self, parent, column_label_mode="number"):
        self.parent = parent
        self.column_label_mode = column_label_mode
        tk.Label(
            parent,
            text="시트 미리보기",
            font=("맑은 고딕", 10, "bold"),
            bg="white",
        ).pack(anchor="w", padx=8, pady=(8, 2))
        self.status_var = tk.StringVar(value="워크북을 불러오는 중입니다.")
        tk.Label(
            parent,
            textvariable=self.status_var,
            bg="white",
            fg="#555555",
            anchor="w",
        ).pack(fill=tk.X, padx=8, pady=(0, 6))
        frame = tk.Frame(parent, bg="white")
        frame.pack(expand=True, fill=tk.BOTH, padx=8, pady=(0, 8))
        self.tree = ttk.Treeview(frame, show="headings")
        y_scroll = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=self.tree.yview)
        x_scroll = ttk.Scrollbar(frame, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(
            yscrollcommand=y_scroll.set,
            xscrollcommand=x_scroll.set,
        )
        self.tree.grid(row=0, column=0, sticky="nsew")
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll.grid(row=1, column=0, sticky="ew")
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)
        self.tree.tag_configure("start", background="#D9EAF7")
        self.tree.tag_configure("header", background="#FCE4C5")

    def clear(self):
        children = self.tree.get_children()
        if children:
            self.tree.delete(*children)

    def _column_heading(self, index):
        if self.column_label_mode == "letter":
            return get_column_letter(index)
        if self.column_label_mode == "letter_number":
            return f"{get_column_letter(index)} ({index})"
        return str(index)

    def refresh(self, worksheet, sheet_name, start, end=None):
        """지정 워크시트의 S/E 주변을 다시 표시하고 실제 표시 범위를 반환합니다."""
        low, high = preview_bounds(start, end)
        last_column = max(1, worksheet.max_column)
        display_columns = min(last_column, MAX_PREVIEW_COLUMNS)
        merged_values = build_merged_value_lookup(
            worksheet,
            minimum_row=low,
            maximum_row=high,
            maximum_column=display_columns,
        )

        columns = ["row"] + [
            f"c{index}" for index in range(1, display_columns + 1)
        ]
        self.clear()
        self.tree["columns"] = columns
        self.tree.heading("row", text="행")
        self.tree.column(
            "row",
            width=52,
            minwidth=52,
            anchor="center",
            stretch=False,
        )
        self.tree.update_idletasks()
        available_width = max(0, self.tree.winfo_width() - 52 - 4)
        data_column_width = max(
            95,
            available_width // display_columns if display_columns else 95,
        )
        for index in range(1, display_columns + 1):
            column_id = f"c{index}"
            self.tree.heading(column_id, text=self._column_heading(index))
            self.tree.column(
                column_id,
                width=data_column_width,
                minwidth=55,
                stretch=True,
            )

        for row_number in range(low, high + 1):
            values = [row_number]
            for column in range(1, display_columns + 1):
                value = worksheet.cell(row=row_number, column=column).value
                if value is None:
                    value = merged_values.get((row_number, column))
                values.append("" if value is None else str(value))
            tag = ""
            if end is not None and start <= row_number <= end:
                tag = "header"
            elif end is None and row_number == start:
                tag = "start"
            self.tree.insert(
                "",
                tk.END,
                values=values,
                tags=(tag,) if tag else (),
            )

        suffix = (
            f" (앞쪽 {MAX_PREVIEW_COLUMNS}열만 표시)"
            if last_column > MAX_PREVIEW_COLUMNS else ""
        )
        if end is None:
            self.status_var.set(
                f"{sheet_name}: {low}~{high}행, S={start} 파란색 표시{suffix}"
            )
        else:
            self.status_var.set(
                f"{sheet_name}: {low}~{high}행, "
                f"S~E={start}~{end} 주황색 표시{suffix}"
            )
        return low, high
