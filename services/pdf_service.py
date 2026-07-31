"""저장된 PDF 매핑 프로필을 사용하는 비GUI PDF/OCR 취합 서비스.

텍스트/OCR 추출기는 기존 엔진의 검증된 구현을 재사용한다. 시각 선택창, 파일 선택창,
알림창은 이 서비스 경로에서 호출하지 않는다.
"""

from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Callable, Mapping, Sequence

import fitz
from openpyxl import load_workbook

import engine_Drag as drag_engine
from core.models import JobResult, JobState, ProgressEvent
from processing_cancellation import ProcessingCancellation
from processing_time import ProcessingTimeEstimator


ProgressReporter = Callable[[ProgressEvent], None]


def _mapping_for(headers: list[str], mapping_set: dict) -> dict:
    fields = {field.get("column"): field for field in mapping_set.get("fields", [])}
    mapping = {}
    for column, header in enumerate(headers, start=1):
        if not header:
            continue
        field = fields.get(column)
        if field is None:
            raise ValueError(f"프로필에 {column}열 '{header}'의 PDF 매핑이 없습니다.")
        mapping[header] = drag_engine._mapping_from_json(field)
    return mapping


def merge_pdfs(
    template_path: str,
    pdfs_by_sheet: Mapping[str, Sequence[str]],
    profile_path: str,
    output_dir: str,
    force_ocr: bool = False,
    report: ProgressReporter | None = None,
    cancellation: ProcessingCancellation | None = None,
) -> JobResult:
    template = Path(template_path)
    if template.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("템플릿은 .xlsx 또는 .xlsm 파일이어야 합니다.")
    if force_ocr and drag_engine.RapidOCR is None:
        raise RuntimeError("강제 OCR에 필요한 rapidocr/onnxruntime을 사용할 수 없습니다.")
    with Path(profile_path).open(encoding="utf-8") as file:
        profile = json.load(file)
    mapping_sets = profile.get("mapping_sets")
    if not isinstance(mapping_sets, list) or not mapping_sets:
        raise ValueError("PDF 매핑 프로필에 mapping_sets가 없습니다.")

    keep_vba = template.suffix.lower() == ".xlsm"
    output_root = Path(output_dir)
    output_root.mkdir(parents=True, exist_ok=True)
    data_book = load_workbook(template, data_only=True, keep_vba=keep_vba)
    write_book = load_workbook(template, keep_vba=keep_vba)
    failed_files: list[str] = []
    summary = {"processed_pages": 0, "empty_pages": [], "rotated_pages": [], "ocr_unavailable_pages": []}
    total_pages = 0
    try:
        for paths in pdfs_by_sheet.values():
            for path in paths:
                if cancellation is not None and cancellation.should_cancel():
                    return JobResult(JobState.CANCELLED, message="PDF merge cancelled")
                with fitz.open(path) as document:
                    total_pages += document.page_count
        overall = {"current": 0}

        planned_work = []
        for mapping_set in mapping_sets:
            weight = 1 + len(mapping_set.get("fields", []))
            for sheet_name in mapping_set.get("sheets", []):
                for path in pdfs_by_sheet.get(sheet_name, []):
                    try:
                        with fitz.open(path) as document:
                            for page in document:
                                if int(page.rotation) % 360:
                                    planned_work.append(("skipped", 1))
                                else:
                                    try:
                                        has_text = bool(page.get_text("text").strip())
                                    except Exception:
                                        has_text = False
                                    planned_work.append(("ocr" if force_ocr or not has_text else "native_text", weight))
                    except Exception:
                        planned_work.append(("failed", 1))
        estimator = ProcessingTimeEstimator(planned_work)

        class CallbackProgress:
            def __init__(self):
                self.current_file = None
                self.current_sheet = None

            def begin_unit(self, file_name, overall_index, detail="", work_type="unknown", ocr_weight=1, sheet_name=""):
                self.current_file = file_name
                self.current_sheet = sheet_name or None
                estimator.begin(work_type, ocr_weight)
                if report:
                    report(ProgressEvent(
                        overall_index - 1, total_pages, detail or "PDF text extraction",
                        file_name, current_sheet=sheet_name or None, activity=work_type,
                        **estimator.metadata(),
                    ))

            def complete_unit(self, completed_units, work_type=None, ocr_weight=None, ocr_initialization_seconds=0):
                estimator.complete(work_type=work_type, weight=ocr_weight, ocr_initialization_seconds=ocr_initialization_seconds)
                if report:
                    metadata = estimator.metadata()
                    report(ProgressEvent(
                        completed_units, total_pages, "PDF page completed",
                        self.current_file, current_sheet=self.current_sheet,
                        activity=work_type, **metadata
                    ))

        any_data = False
        for mapping_set in mapping_sets:
            sheets = mapping_set.get("sheets")
            start, end = mapping_set.get("header_start"), mapping_set.get("header_end")
            if not isinstance(sheets, list) or not sheets or not isinstance(start, int) or not isinstance(end, int):
                raise ValueError("유효하지 않은 PDF 매핑 세트입니다.")
            master_sheet = sheets[0]
            if master_sheet not in data_book.sheetnames:
                raise ValueError(f"템플릿에 없는 시트입니다: {master_sheet}")
            headers = drag_engine._extract_headers(data_book[master_sheet], start, end)
            if not any(headers):
                raise ValueError(f"{master_sheet}에서 유효한 헤더를 찾지 못했습니다.")
            mapping = _mapping_for(headers, mapping_set)
            for sheet_name in sheets:
                if cancellation is not None and cancellation.should_cancel():
                    return JobResult(JobState.CANCELLED, message="PDF merge cancelled")
                paths = [str(Path(path)) for path in pdfs_by_sheet.get(sheet_name, [])]
                if not paths:
                    raise ValueError(f"{sheet_name} 시트에 연결된 PDF 파일이 없습니다.")
                rows = drag_engine._collect_pdf_rows(
                    paths, headers, mapping, force_ocr, failed_files, summary,
                    CallbackProgress(), overall, cancellation, sheet_name
                )
                if cancellation is not None and cancellation.should_cancel():
                    return JobResult(JobState.CANCELLED, message="PDF merge cancelled")
                if not rows:
                    continue
                any_data = True
                worksheet = write_book[sheet_name]
                merged = drag_engine._build_merged_write_lookup(worksheet)
                for offset, values in enumerate(rows, start=1):
                    output_row = end + offset
                    for column, header in enumerate(headers, start=1):
                        if not header:
                            continue
                        target = merged.get((output_row, column))
                        if target is not None and target != (output_row, column):
                            continue
                        worksheet.cell(row=output_row, column=column, value=values.get(header, ""))
        if not any_data:
            return JobResult(JobState.FAILED, failed_files=failed_files, message="PDF에서 추출할 데이터가 없습니다.", details=summary)
        extension = ".xlsm" if keep_vba else ".xlsx"
        output = output_root / f"DnS_Auto_Drag취합_{datetime.now():%H%M%S_%f}{extension}"
        if cancellation is not None:
            cancellation.reserve_output(output)
            if not cancellation.enter_save_phase():
                cancellation.rollback_outputs()
                return JobResult(JobState.CANCELLED, message="PDF merge cancelled")
        write_book.save(output)
        return JobResult(
            JobState.SUCCEEDED if not failed_files else JobState.FAILED,
            output_files=[str(output)], failed_files=failed_files, message="PDF 취합 완료", details=summary,
        )
    finally:
        data_book.close()
        write_book.close()
        drag_engine.TEXT_EXTRACTOR.reset_work_cache()
