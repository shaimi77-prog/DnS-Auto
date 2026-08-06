"""저장된 시트 설정 프로필을 이용하는 비GUI Excel 취합 서비스."""

from __future__ import annotations

import json
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Callable, Iterable

from openpyxl import load_workbook

from core.models import JobResult, JobState, ProgressEvent
from processing_cancellation import ProcessingCancellation
from processing_time import ProcessingTimeEstimator
from com_process_ownership import Ownership, capture_processes, cleanup_com_session, confirm_ownership, detached_quit_callback


ProgressReporter = Callable[[ProgressEvent], None]


def _convert_xls_to_xlsx(source: Path) -> Path:
    """Excel COM을 이용해 구형 XLS를 임시 XLSX로 변환한다."""
    import pythoncom
    import win32com.client as win32

    handle = tempfile.NamedTemporaryFile(prefix="dns-auto-xls-", suffix=".xlsx", delete=False)
    converted = Path(handle.name)
    handle.close()
    converted.unlink(missing_ok=True)
    excel = workbook = None
    ownership = Ownership("unconfirmed")
    try:
        before = capture_processes(["EXCEL.EXE"])
    except Exception:
        before = {}
    pythoncom.CoInitialize()
    try:
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        try:
            ownership = confirm_ownership(before, int(excel.Hwnd), ["EXCEL.EXE"])
        except Exception:
            ownership = Ownership("unconfirmed")
        workbook = excel.Workbooks.Open(
            str(source.resolve()),
            Password="DummyPassword123!",
            IgnoreReadOnlyRecommended=True,
        )
        workbook.SaveAs(str(converted.resolve()), FileFormat=51)
        return converted
    except Exception:
        converted.unlink(missing_ok=True)
        raise
    finally:
        if workbook is not None:
            try:
                workbook.Close(SaveChanges=False)
            except Exception:
                pass
            workbook = None
        quit_callback = detached_quit_callback(excel) if excel is not None else None
        excel = None
        cleanup_com_session(
            application="excel",
            close_callbacks=[],
            quit_callback=quit_callback,
            ownership=ownership,
            co_uninitialize=pythoncom.CoUninitialize,
            allow_forced_cleanup=True,
        )


def _last_data_row(worksheet, minimum_row: int) -> int:
    for row in range(worksheet.max_row, minimum_row - 1, -1):
        if any(cell.value not in (None, "") for cell in worksheet[row]):
            return row
    return minimum_row


def _load_configs(profile_path: str) -> dict[str, dict]:
    with Path(profile_path).open(encoding="utf-8") as file:
        profile = json.load(file)
    if profile.get("profile_type") != "sheet_config":
        raise ValueError("Sheet 설정 프로필이 아닙니다.")
    configs = profile.get("sheet_configs")
    if not isinstance(configs, list) or not configs:
        raise ValueError("sheet_configs가 없거나 비어 있습니다.")
    normalized = {}
    for config in configs:
        name = config.get("sheet_name")
        start, end = config.get("header_start"), config.get("header_end")
        if (
            not isinstance(name, str)
            or not isinstance(start, int)
            or not isinstance(end, int)
            or start < 1
            or end < start
        ):
            raise ValueError("유효하지 않은 시트 설정 프로필입니다.")
        normalized[name] = {
            "S": start,
            "E": end,
            "mode": int(config.get("mode", 1)),
            "key_col": str(config.get("key_col") or ""),
            "protect": bool(config.get("protect", True)),
        }
    return normalized


def merge_workbooks(
    template_path: str,
    source_paths: Iterable[str],
    profile_path: str | None,
    output_dir: str,
    report: ProgressReporter | None = None,
    settings: dict[str, dict] | None = None,
    cancellation: ProcessingCancellation | None = None,
) -> JobResult:
    """기존 Sheet 엔진의 행 복사 규칙을 GUI 없이 실행한다."""
    template = Path(template_path)
    sources = [Path(path) for path in source_paths]
    if template.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("템플릿은 .xlsx 또는 .xlsm 파일이어야 합니다.")
    if not sources:
        raise ValueError("취합할 원본 파일이 없습니다.")
    configs = settings if settings is not None else _load_configs(profile_path or "")
    if not configs:
        raise ValueError("Sheet 설정이 없습니다.")

    keep_vba = template.suffix.lower() == ".xlsm"
    output_root = Path(output_dir)
    output_root.mkdir(parents=True, exist_ok=True)
    target = load_workbook(template, keep_vba=keep_vba)
    failed, copied_rows = [], 0
    work_types = ["xls_merge" if source.suffix.lower() == ".xls" else "xlsx_merge" for source in sources]
    estimator = ProcessingTimeEstimator(zip(work_types, [1] * len(sources)), minimum_samples=1)
    estimator.start()

    def event(completed, message, source, *, sheet=None):
        return ProgressEvent(
            completed, len(sources), message, str(source), current_sheet=sheet,
            activity="excel", **estimator.metadata()
        )
    try:
        missing_sheets = set(configs) - set(target.sheetnames)
        if missing_sheets:
            raise ValueError(
                f"템플릿에 없는 시트가 프로필에 있습니다: {', '.join(sorted(missing_sheets))}"
            )
        states = {}
        for name, config in configs.items():
            worksheet = target[name]
            state = {}
            if config["mode"] == 1:
                state["append_row"] = max(
                    config["E"], _last_data_row(worksheet, config["E"])
                ) + 1
            elif config["mode"] == 2 and config["key_col"]:
                state["row_map"] = {
                    str(worksheet[f"{config['key_col']}{row}"].value).strip(): row
                    for row in range(config["E"] + 1, worksheet.max_row + 1)
                    if worksheet[f"{config['key_col']}{row}"].value not in (None, "")
                }
            states[name] = state

        for index, source in enumerate(sources, start=1):
            if cancellation is not None and cancellation.should_cancel():
                return JobResult(
                    JobState.CANCELLED,
                    message="Excel 취합이 취소되었습니다.",
                    details={"timing": estimator.summary()},
                )
            if report:
                report(
                    event(index - 1, "Excel 취합 중", source)
                )
            if source.resolve() == template.resolve():
                failed.append(str(source))
                estimator.begin(work_types[index - 1])
                estimator.complete(successful=False)
                if report:
                    report(event(index, "Excel 취합 건너뜀", source))
                continue
            estimator.begin(work_types[index - 1])
            workbook = None
            converted = None
            successful = False
            try:
                source_to_open = source
                if source.suffix.lower() == ".xls":
                    converted = _convert_xls_to_xlsx(source)
                    source_to_open = converted
                workbook = load_workbook(source_to_open, data_only=True)
                for name, config in configs.items():
                    if report:
                        report(
                            event(index - 1, "Excel 시트 취합 중", source, sheet=name)
                        )
                    if name not in workbook.sheetnames:
                        continue
                    source_sheet = workbook[name]
                    target_sheet = target[name]
                    state = states[name]
                    empty_count = 0
                    for row_number in range(config["E"] + 1, source_sheet.max_row + 1):
                        if source_sheet.row_dimensions[row_number].hidden:
                            continue
                        values = {
                            column: source_sheet.cell(row=row_number, column=column).value
                            for column in range(1, source_sheet.max_column + 1)
                            if source_sheet.cell(row=row_number, column=column).value
                            not in (None, "")
                        }
                        if not values:
                            empty_count += 1
                            if empty_count >= 200:
                                break
                            continue
                        empty_count = 0
                        target_row = None
                        if config["mode"] == 1:
                            target_row = state["append_row"]
                            state["append_row"] += 1
                        elif config["key_col"]:
                            key = source_sheet[
                                f"{config['key_col']}{row_number}"
                            ].value
                            if key not in (None, ""):
                                target_row = state["row_map"].get(str(key).strip())
                        else:
                            target_row = row_number
                        if target_row is None:
                            continue
                        wrote = False
                        for column, value in values.items():
                            cell = target_sheet.cell(row=target_row, column=column)
                            if type(cell).__name__ == "MergedCell":
                                continue
                            if config["protect"] and cell.value not in (None, ""):
                                continue
                            cell.value = value
                            wrote = True
                        copied_rows += int(wrote)
                successful = True
            except Exception:
                failed.append(str(source))
            finally:
                if workbook is not None:
                    workbook.close()
                if converted is not None:
                    converted.unlink(missing_ok=True)
            sampled = successful and not (
                cancellation is not None and cancellation.should_cancel()
            )
            estimator.complete(successful=sampled)
            if report:
                report(event(index, "Excel 취합 완료", source))
            if cancellation is not None and cancellation.should_cancel():
                if sampled:
                    estimator.discard_last_sample(work_types[index - 1])
                return JobResult(
                    JobState.CANCELLED,
                    message="Excel 취합이 취소되었습니다.",
                    details={"timing": estimator.summary()},
                )

        if copied_rows == 0:
            return JobResult(
                JobState.FAILED,
                failed_files=failed,
                message="취합할 데이터가 없습니다.",
                details={"timing": estimator.summary()},
            )
        extension = ".xlsm" if keep_vba else ".xlsx"
        output = output_root / (
            f"DnS_Auto_Sheet다중취합_{datetime.now():%H%M%S_%f}{extension}"
        )
        if cancellation is not None and not cancellation.enter_save_phase():
            return JobResult(
                JobState.CANCELLED,
                message="Excel 취합이 취소되었습니다.",
                details={"timing": estimator.summary()},
            )
        target.save(output)
        return JobResult(
            JobState.SUCCEEDED if not failed else JobState.FAILED,
            output_files=[str(output)],
            failed_files=failed,
            message="Excel 취합 완료",
            details={"copied_rows": copied_rows, "timing": estimator.summary()},
        )
    finally:
        target.close()
