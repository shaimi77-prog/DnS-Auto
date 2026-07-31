"""Discover portable inputs and profiles without making an arbitrary merge choice."""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any

EXCEL_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}
PDF_EXTENSIONS = {".pdf"}
TEMPLATE_HINTS = ("양식", "template", "서식", "form")


def _display(path: Path, app_root: Path) -> str:
    try:
        return str(path.resolve().relative_to(app_root.resolve()))
    except ValueError:
        return str(path.resolve())


def _workbook_sheets(path: Path) -> set[str] | None:
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return None
    try:
        from openpyxl import load_workbook
        workbook = load_workbook(path, read_only=True, data_only=False)
        try:
            return set(workbook.sheetnames)
        finally:
            workbook.close()
    except Exception:
        return None


def _profile_sheets(profile: dict[str, Any], profile_type: str) -> set[str]:
    if profile_type == "sheet_config":
        return {str(item.get("sheet_name")) for item in profile.get("sheet_configs", []) if item.get("sheet_name")}
    sheets: set[str] = set()
    for item in profile.get("mapping_sets", []):
        if isinstance(item, dict):
            sheets.update(str(value) for value in item.get("sheets", []) if value)
    return sheets


def _profiles(app_root: Path, operation: str, template: Path | None) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    expected_type = "sheet_config" if operation == "excel" else "pdf_mapping"
    folder = app_root / "profiles" / ("sheet" if operation == "excel" else "pdf")
    compatible: list[dict[str, Any]] = []
    rejected: list[dict[str, Any]] = []
    template_sheets = _workbook_sheets(template) if template else None
    for path in sorted(folder.rglob("*.json")) if folder.is_dir() else []:
        try:
            profile = json.loads(path.read_text(encoding="utf-8"))
            actual_type = profile.get("profile_type")
            legacy_pdf = operation == "pdf" and actual_type is None and "mapping_sets" in profile
            if actual_type != expected_type and not legacy_pdf:
                raise ValueError(f"프로필 유형이 {expected_type}이 아닙니다.")
            metadata = profile.get("metadata") if isinstance(profile.get("metadata"), dict) else {}
            expected_name = metadata.get("template_file_name")
            reasons: list[str] = []
            if template and expected_name and expected_name != template.name:
                reasons.append(f"기준 양식명이 다릅니다: {expected_name}")
            required_sheets = _profile_sheets(profile, expected_type)
            if template_sheets is not None and required_sheets - template_sheets:
                reasons.append("양식에 없는 시트가 필요합니다: " + ", ".join(sorted(required_sheets - template_sheets)))
            item = {
                "name": path.stem,
                "profile_name": metadata.get("profile_name") or path.stem,
                "path": _display(path, app_root),
                "profile_type": expected_type,
                "template_file_name": expected_name,
                "sheets": sorted(required_sheets),
                "compatibility": "incompatible" if reasons else ("verified" if template else "type_only"),
            }
            if reasons:
                item["reasons"] = reasons
                rejected.append(item)
            else:
                compatible.append(item)
        except (OSError, ValueError, json.JSONDecodeError, TypeError) as error:
            rejected.append({"name": path.stem, "path": _display(path, app_root), "compatibility": "invalid", "reasons": [str(error)]})
    return compatible, rejected


def discover_merge_plan(
    app_root: str | Path,
    output_root: str | Path,
    input_root: str | Path | None = None,
    operation: str = "auto",
    template_path: str | None = None,
    profile_name: str | None = None,
    interactive: bool = False,
) -> dict[str, Any]:
    app = Path(app_root).resolve()
    inputs = Path(input_root).resolve() if input_root else (app / "inputs").resolve()
    output = Path(output_root).resolve()
    if operation not in {"auto", "excel", "pdf"}:
        raise ValueError("operation은 auto, excel 또는 pdf여야 합니다.")
    if not inputs.is_dir():
        return {"status": "needs_clarification", "operation": None, "input_root": str(inputs), "output_root": str(output), "questions": [f"입력 폴더를 찾을 수 없습니다. 사용할 경로를 알려주세요: {inputs}"]}

    files = sorted(path.resolve() for path in inputs.rglob("*") if path.is_file())
    excel_files = [path for path in files if path.suffix.lower() in EXCEL_EXTENSIONS and not path.name.startswith("~$")]
    pdf_files = [path for path in files if path.suffix.lower() in PDF_EXTENSIONS]
    explicit_template = Path(template_path).resolve() if template_path else None
    if explicit_template and not explicit_template.is_file():
        raise ValueError(f"지정한 양식을 찾을 수 없습니다: {explicit_template}")

    detected = operation
    if operation == "auto":
        if pdf_files and excel_files:
            detected = "pdf"
        elif len(excel_files) >= 2:
            detected = "excel"
        else:
            detected = "unknown"

    if explicit_template:
        template_candidates = [explicit_template]
    else:
        hinted = [path for path in excel_files if any(hint in path.stem.lower() for hint in TEMPLATE_HINTS)]
        template_candidates = hinted if hinted else excel_files
    selected_template = template_candidates[0] if len(template_candidates) == 1 else None

    if detected == "excel":
        sources = [path for path in excel_files if path != selected_template]
    elif detected == "pdf":
        sources = pdf_files
    else:
        sources = files

    compatible: list[dict[str, Any]] = []
    rejected: list[dict[str, Any]] = []
    if detected in {"excel", "pdf"}:
        compatible, rejected = _profiles(app, detected, selected_template)

    selected_profile = None
    profile_matches: list[dict[str, Any]] = []
    if profile_name:
        wanted = Path(profile_name).stem.casefold()
        profile_matches = [item for item in compatible if item["name"].casefold() == wanted or str(item["profile_name"]).casefold() == wanted]
        if len(profile_matches) == 1:
            selected_profile = profile_matches[0]

    questions: list[str] = []
    if detected == "unknown":
        questions.append("Excel 취합인지 PDF 취합인지 판단할 수 없습니다. 작업 종류와 입력 파일을 알려주세요.")
    if not template_candidates:
        questions.append("기준 Excel 양식을 찾지 못했습니다. 양식 경로를 알려주세요.")
    elif len(template_candidates) > 1:
        questions.append("기준 양식 후보가 여러 개입니다. 사용할 양식을 선택해 주세요: " + ", ".join(_display(path, app) for path in template_candidates))
    if detected in {"excel", "pdf"} and not sources:
        questions.append("취합할 원본 파일을 찾지 못했습니다. 입력 폴더 또는 파일 경로를 알려주세요.")
    if profile_name and len(profile_matches) != 1:
        questions.append(f"'{profile_name}'과 정확히 일치하는 호환 프로필을 하나 찾지 못했습니다. 프로필명을 확인해 주세요.")

    profile_confirmation = False
    if not questions and not interactive and not profile_name:
        if compatible:
            profile_confirmation = True
            questions.append("호환 가능한 프로필을 발견했습니다. 사용할 프로필을 지정하거나 대화형 진행을 선택해 주세요: " + ", ".join(item["name"] for item in compatible))
        else:
            profile_confirmation = True
            questions.append("호환 가능한 저장 프로필이 없습니다. 대화형으로 설정 창을 열어 진행할까요?")

    if questions:
        status = "needs_confirmation" if profile_confirmation and len(questions) == 1 else "needs_clarification"
    else:
        status = "ready"

    result: dict[str, Any] = {
        "status": status,
        "operation": None if detected == "unknown" else detected,
        "input_root": _display(inputs, app),
        "output_root": _display(output, app),
        "template_candidates": [_display(path, app) for path in template_candidates],
        "selected_template": _display(selected_template, app) if selected_template else None,
        "source_candidates": [_display(path, app) for path in sources],
        "compatible_profiles": compatible,
        "rejected_profiles": rejected,
        "selected_profile": selected_profile,
        "interactive": bool(interactive),
        "questions": questions,
    }
    if status == "ready":
        if interactive:
            tool = "start_interactive_sheet_merge" if detected == "excel" else "start_interactive_pdf_merge"
        else:
            tool = "start_sheet_merge" if detected == "excel" else "start_pdf_merge"
        result["next_tool"] = tool
        result["instructions"] = "확인 질문 없이 next_tool을 호출해 진행하세요. 실행 전 반환된 경로를 그대로 사용하세요."
    else:
        result["instructions"] = "임의로 후보를 선택하거나 작업을 시작하지 말고 questions를 사용자에게 확인하세요."
    return result