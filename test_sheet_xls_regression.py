import os
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

import engine_Sheet
from services import sheet_service


class SheetXlsRegressionTests(unittest.TestCase):
    def test_converts_xls_instead_of_silently_skipping_it(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            template = root / "template.xlsx"
            legacy = root / "agency.xls"
            converted = root / "converted.xlsx"
            output = root / "outputs"
            legacy.write_bytes(b"legacy placeholder")

            book = Workbook()
            sheet = book.active
            sheet.title = "Data"
            sheet.append(["id", "agency", "value"])
            sheet.append([None, "기관A", None])
            book.save(template)

            book = Workbook()
            sheet = book.active
            sheet.title = "Data"
            sheet.append(["id", "agency", "value"])
            sheet.append([1, "기관A", 123])
            book.save(converted)

            settings = {
                "Data": {
                    "S": 1, "E": 1, "mode": 2,
                    "key_col": "B", "protect": True,
                }
            }
            with patch.object(
                sheet_service, "_convert_xls_to_xlsx", return_value=converted
            ) as convert:
                result = sheet_service.merge_workbooks(
                    str(template), [str(legacy)], None, str(output),
                    settings=settings,
                )
            convert.assert_called_once_with(legacy)
            self.assertEqual(result.failed_files, [])
            result_book = load_workbook(result.output_files[0], data_only=True)
            self.assertEqual(result_book["Data"]["C2"].value, 123)

    def test_opens_result_directory(self):
        output = os.path.abspath(os.path.join("result", "output.xlsx"))
        with patch.object(engine_Sheet.os, "startfile", create=True) as startfile:
            self.assertTrue(engine_Sheet._open_result_folder(output))
        startfile.assert_called_once_with(os.path.dirname(output))

    def test_folder_open_failure_does_not_turn_success_into_error(self):
        with patch.object(
            engine_Sheet.os, "startfile", side_effect=OSError("blocked"),
            create=True,
        ):
            self.assertFalse(engine_Sheet._open_result_folder("output.xlsx"))


if __name__ == "__main__":
    unittest.main()
