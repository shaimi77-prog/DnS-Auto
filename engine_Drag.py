"""PDF 영역 지정, 앵커 추적 및 OCR 기반 데이터 취합 기능."""
# Copyright (C) 2026 두부코드(DOOBOO_CODE)
# SPDX-License-Identifier: AGPL-3.0-only

import datetime
import logging
import os
import queue
import sys
import threading
import time
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

import fitz  # PyMuPDF
import numpy as np
from openpyxl import load_workbook
from PIL import Image, ImageTk

from config import PROGRAM_NAME, VERSION
from page_preprocessing import (
    DocumentOrientationClassifier,
    PagePreprocessConfig,
    PagePreprocessor,
    inverse_transform_box,
)
from reference_navigation import ReferenceNavigationState
from utils_profiles import (
    PDF_PROFILE_TYPE,
    PROFILE_SCHEMA_VERSION,
    application_dir,
    prepare_profile_directory,
    profile_directory,
    read_profile,
    write_profile,
)
from utils_progress import ProcessingProgressDialog
from utils_sheet_preview import (
    EXCEL_MAX_ROW,
    HEADER_PREVIEW_DEBOUNCE_MS,
    HEADER_PREVIEW_MARGIN,
    SheetPreviewPanel,
    build_merged_value_lookup as _build_merged_value_lookup,
    preview_bounds as _preview_bounds,
)

MAX_ANCHOR_KEYWORD_LENGTH = 10
GROUP_OPTIONS = [f"그룹 {index}" for index in range(1, 6)] + ["개별"]
REFERENCE_SCAN_MAX_PAGES = 3
MIN_NATIVE_TEXT_CHARS = 30
MIN_NATIVE_TEXT_BLOCKS = 3
OCR_KEYWORD_SEARCH_DPI = 150
OCR_VALUE_RECOGNITION_DPI = 180
OCR_MIN_TEXT_SCORE = 0.4
OCR_CANDIDATE_MERGE_DISTANCE_PT = 5.0
OCR_CANDIDATE_IOU_THRESHOLD = 0.6

NATIVE_MATCH = "NATIVE_MATCH"
OCR_MATCH = "OCR_MATCH"
OCR_NO_MATCH = "OCR_NO_MATCH"
OCR_UNAVAILABLE = "OCR_UNAVAILABLE"
ROTATED_PAGE_EXCLUDED = "ROTATED_PAGE_EXCLUDED"


def validate_anchor_keyword(keyword):
    """앞뒤 공백만 제거하고 중간 공백을 보존한 채 길이를 검증합니다."""
    cleaned = (keyword or "").strip()
    return cleaned, len(cleaned) <= MAX_ANCHOR_KEYWORD_LENGTH


try:
    from rapidocr import LangRec, ModelType, OCRVersion, RapidOCR
except (ImportError, OSError) as ocr_import_error:
    OCR_IMPORT_ERROR = str(ocr_import_error)
    RapidOCR = None
    LangRec = ModelType = OCRVersion = None
    logging.warning(f"OCR 모듈을 불러오지 못했습니다. Native PDF 모드로 계속합니다: {ocr_import_error}")
else:
    OCR_IMPORT_ERROR = None


class HybridTextExtractor:
    """Native PDF 텍스트와 로컬 ONNX OCR을 선택적으로 사용하는 단일 추출기."""
    _instance = None

    def __new__(cls):
        if cls._instance is None:
            cls._instance = super().__new__(cls)
        return cls._instance

    def __init__(self):
        if hasattr(self, "_initialized"):
            return
        self._initialized = True
        self.ocr = None
        self.ocr_unavailable_reason = None
        self._ocr_unavailable_logged = False
        self._ocr_disabled_for_work = False
        self._ocr_cache = {}
        self.last_keyword_status = None
        self.last_keyword_reason = None
        self.last_ocr_initialization_seconds = 0.0
        self.page_preprocessor = PagePreprocessor(
            PagePreprocessConfig(
                enabled=True,
                orientation_enabled=True,
                deskew_enabled=True,
            ),
            orientation_classifier=DocumentOrientationClassifier(),
        )
        self.last_preprocess_result = None

    def reset_work_cache(self):
        """새 취합 작업을 시작할 때 메모리 OCR 캐시와 상태를 초기화합니다."""
        self._ocr_cache.clear()
        self.last_keyword_status = None
        self.last_keyword_reason = None
        self.last_ocr_initialization_seconds = 0.0
        self.last_preprocess_result = None
        self._ocr_unavailable_logged = False
        self._ocr_disabled_for_work = False
        if self.ocr is not None:
            self.ocr_unavailable_reason = None

    def release_pdf_cache(self, pdf_path):
        """처리가 끝난 PDF의 OCR 결과를 즉시 해제합니다."""
        normalized = os.path.abspath(pdf_path)
        self._ocr_cache = {
            key: value
            for key, value in self._ocr_cache.items()
            if key[0] != normalized
        }

    def _set_ocr_unavailable(self, reason, detail):
        self.ocr_unavailable_reason = reason
        self._ocr_disabled_for_work = True
        if not self._ocr_unavailable_logged:
            logging.error(
                "%s: reason=%s, detail=%s",
                OCR_UNAVAILABLE,
                reason,
                detail,
            )
            self._ocr_unavailable_logged = True

    def ensure_ocr(self):
        """OCR이 실제로 필요할 때만 무거운 ONNX 모델을 초기화합니다."""
        if self._ocr_disabled_for_work:
            return False
        if self.ocr is not None:
            return True
        if RapidOCR is None:
            self._set_ocr_unavailable(
                "MODULE_IMPORT_FAILED",
                "RapidOCR 모듈을 불러오지 못했습니다."
                + (f" 사유: {OCR_IMPORT_ERROR}" if OCR_IMPORT_ERROR else ""),
            )
            return False
        initialization_started_at = time.monotonic()
        try:
            base_path = sys._MEIPASS if getattr(sys, "frozen", False) else os.path.dirname(os.path.abspath(__file__))
            model_dir = os.path.join(base_path, "ocr_models")
            model_paths = {
                "Det.model_path": os.path.join(model_dir, "ch_PP-OCRv5_det_mobile.onnx"),
                "Cls.model_path": os.path.join(model_dir, "ch_ppocr_mobile_v2.0_cls_mobile.onnx"),
                "Rec.model_path": os.path.join(model_dir, "korean_PP-OCRv5_rec_mobile.onnx"),
            }
            missing = [path for path in model_paths.values() if not os.path.isfile(path)]
            if missing:
                self._set_ocr_unavailable(
                    "MODEL_FILE_MISSING",
                    f"내장 OCR 모델 파일이 누락되었습니다: {missing}",
                )
                return False
            self.ocr = RapidOCR(params={
                **model_paths,
                "Rec.lang_type": LangRec.KOREAN,
                "Rec.ocr_version": OCRVersion.PPOCRV5,
                "Rec.model_type": ModelType.MOBILE,
                "Global.log_level": "warning",
            })
            self.last_ocr_initialization_seconds = time.monotonic() - initialization_started_at
            logging.info("한국어 PP-OCRv5 로컬 ONNX 엔진 초기화 완료")
            self.ocr_unavailable_reason = None
            return True
        except Exception as error:
            self._set_ocr_unavailable("ENGINE_INITIALIZATION_FAILED", str(error))
            logging.debug("ONNX OCR 엔진 초기화 예외", exc_info=True)
            return False

    @staticmethod
    def _distance(rect1, rect2):
        return (
            ((rect1.x0 + rect1.x1 - rect2.x0 - rect2.x1) / 2) ** 2
            + ((rect1.y0 + rect1.y1 - rect2.y0 - rect2.y1) / 2) ** 2
        ) ** 0.5

    @staticmethod
    def _normalized_keyword(text):
        """OCR 비교용으로 공백만 정규화하며 표시 원문은 변경하지 않습니다."""
        return "".join((text or "").split()).casefold()

    def _keyword_subrect(self, rect, detected_text, keyword):
        """Approximate the keyword portion when OCR returns a longer text box."""
        normalized_text = self._normalized_keyword(detected_text)
        normalized_keyword = self._normalized_keyword(keyword)
        if (
            not normalized_text
            or not normalized_keyword
            or normalized_text == normalized_keyword
        ):
            return fitz.Rect(rect)
        start = normalized_text.find(normalized_keyword)
        if start < 0:
            return fitz.Rect(rect)
        unit_width = rect.width / len(normalized_text)
        return fitz.Rect(
            rect.x0 + unit_width * start,
            rect.y0,
            rect.x0 + unit_width * (start + len(normalized_keyword)),
            rect.y1,
        )

    @staticmethod
    def _intersection_over_union(rect1, rect2):
        intersection = fitz.Rect(rect1) & fitz.Rect(rect2)
        if intersection.is_empty:
            return 0.0
        intersection_area = intersection.width * intersection.height
        union_area = (
            rect1.width * rect1.height
            + rect2.width * rect2.height
            - intersection_area
        )
        return intersection_area / union_area if union_area > 0 else 0.0

    def _merge_ocr_candidates(self, candidates):
        """OCR 단계별 중복 후보를 중심점 거리와 박스 중첩률로 병합합니다."""
        merged = []
        for candidate in sorted(
            candidates,
            key=lambda item: (-item["score"], item["step"]),
        ):
            duplicate_index = None
            for index, existing in enumerate(merged):
                if (
                    self._distance(candidate["rect"], existing["rect"])
                    <= OCR_CANDIDATE_MERGE_DISTANCE_PT
                    or self._intersection_over_union(
                        candidate["rect"],
                        existing["rect"],
                    )
                    >= OCR_CANDIDATE_IOU_THRESHOLD
                ):
                    duplicate_index = index
                    break
            if duplicate_index is None:
                merged.append(candidate)
            elif candidate["score"] > merged[duplicate_index]["score"]:
                merged[duplicate_index] = candidate
        return merged

    @staticmethod
    def _page_identity(page):
        document_name = getattr(page.parent, "name", "") or ""
        return (
            os.path.abspath(document_name)
            if document_name
            else f"<memory:{id(page.parent)}>"
        )

    def _ocr_detect(self, page, clip, dpi):
        """이미지 배열을 제외한 OCR 문자열·점수·박스만 작업 캐시에 저장합니다."""
        key = (
            self._page_identity(page),
            int(page.number),
            tuple(round(value, 2) for value in clip),
            int(dpi),
            True,
        )
        if key in self._ocr_cache:
            return self._ocr_cache[key]
        if not self.ensure_ocr():
            return None
        try:
            pix = page.get_pixmap(clip=clip, dpi=dpi, alpha=False)
        except Exception as error:
            self._set_ocr_unavailable("PAGE_RENDER_FAILED", str(error))
            return None
        transform_matrix = np.eye(3, dtype=np.float64)
        source_width = pix.width
        source_height = pix.height
        ocr_input = pix.tobytes("png")
        preprocess_result = None
        if fitz.Rect(clip) == page.rect:
            rgb = np.frombuffer(pix.samples, dtype=np.uint8).reshape(
                pix.height, pix.width, pix.n
            )[:, :, :3]
            preprocess_result = self.page_preprocessor.preprocess(
                rgb[:, :, ::-1], pdf_rotation=page.rotation
            )
            self.last_preprocess_result = preprocess_result
            transform_matrix = preprocess_result.transform_matrix
            if preprocess_result.status == "corrected":
                ocr_input = preprocess_result.processed_image
            if preprocess_result.orientation_applied:
                logging.info(
                    "PAGE_ORIENTATION_APPLIED: 파일=%s, 페이지=%s, "
                    "감지=%s도, 보정=%s도, 신뢰도=%.4f, 차이=%.4f",
                    os.path.basename(self._page_identity(page)),
                    page.number + 1,
                    preprocess_result.detected_orientation,
                    preprocess_result.orientation_correction,
                    preprocess_result.orientation_confidence,
                    preprocess_result.orientation_margin,
                )
            if preprocess_result.deskew_applied:
                logging.info(
                    "PAGE_DESKEW_APPLIED: 파일=%s, 페이지=%s, 각도=%.3f, "
                    "선분=%s, 분산=%.3f",
                    os.path.basename(self._page_identity(page)),
                    page.number + 1,
                    preprocess_result.skew_angle,
                    preprocess_result.valid_line_count,
                    preprocess_result.angle_dispersion,
                )
        try:
            result = self.ocr(
                ocr_input,
                use_det=True,
                use_cls=False,
                use_rec=True,
            )
        except Exception as error:
            self._set_ocr_unavailable("OCR_EXECUTION_ERROR", str(error))
            logging.debug("OCR 검출 실행 예외", exc_info=True)
            return None
        boxes = getattr(result, "boxes", None)
        texts = tuple(getattr(result, "txts", None) or ())
        scores = tuple(getattr(result, "scores", None) or ())
        compact = {
            "width": (
                preprocess_result.processed_image.shape[1]
                if preprocess_result is not None else pix.width
            ),
            "height": (
                preprocess_result.processed_image.shape[0]
                if preprocess_result is not None else pix.height
            ),
            "source_width": source_width,
            "source_height": source_height,
            "transform_matrix": transform_matrix,
            "preprocess_status": (
                preprocess_result.status
                if preprocess_result is not None else "not_requested"
            ),
            "clip": tuple(clip),
            "boxes": tuple(boxes) if boxes is not None else (),
            "texts": texts,
            "scores": scores,
        }
        self._ocr_cache[key] = compact
        return compact

    def _ocr_page_detect(self, page):
        """Cache one full-page OCR result for keyword tracking and value extraction."""
        return self._ocr_detect(page, page.rect, OCR_KEYWORD_SEARCH_DPI)

    @staticmethod
    def _ocr_box_to_pdf_rect(
        box,
        clip,
        image_width,
        image_height,
        transform_matrix=None,
        source_width=None,
        source_height=None,
    ):
        points = np.asarray(list(box), dtype=np.float64)
        if transform_matrix is not None:
            points = inverse_transform_box(points, transform_matrix)
        xs = [float(point[0]) for point in points]
        ys = [float(point[1]) for point in points]
        coordinate_width = source_width or image_width
        coordinate_height = source_height or image_height
        return fitz.Rect(
            clip.x0 + min(xs) * clip.width / coordinate_width,
            clip.y0 + min(ys) * clip.height / coordinate_height,
            clip.x0 + max(xs) * clip.width / coordinate_width,
            clip.y0 + max(ys) * clip.height / coordinate_height,
        )
    def _find_native_keyword_candidates(self, page, origin_rect, keyword):
        if not keyword:
            return []
        width = max(origin_rect.width, 5)
        height = max(origin_rect.height, 5)
        page_rect = page.rect
        candidates = {}
        for step in range(1, 6):
            clip = fitz.Rect(
                max(page_rect.x0, origin_rect.x0 - width * step / 5),
                max(page_rect.y0, origin_rect.y0 - height * step / 5),
                min(page_rect.x1, origin_rect.x1 + width * step / 5),
                min(page_rect.y1, origin_rect.y1 + height * step / 5),
            )
            for match in page.search_for(keyword, clip=clip):
                key = tuple(
                    round(value, 3)
                    for value in (match.x0, match.y0, match.x1, match.y1)
                )
                candidates.setdefault(key, fitz.Rect(match))
        return list(candidates.values())

    def _find_ocr_keyword_candidates(self, page, origin_rect, keyword):
        if not keyword:
            return []
        normalized_keyword = self._normalized_keyword(keyword)
        width = max(origin_rect.width, 5)
        height = max(origin_rect.height, 5)
        page_rect = page.rect
        candidates = []
        clips = [
            fitz.Rect(
                max(page_rect.x0, origin_rect.x0 - width * step / 5),
                max(page_rect.y0, origin_rect.y0 - height * step / 5),
                min(page_rect.x1, origin_rect.x1 + width * step / 5),
                min(page_rect.y1, origin_rect.y1 + height * step / 5),
            )
            for step in range(1, 6)
        ]
        # 동일 DPI에서는 5차 영역을 한 번 검출해도 1~5차 전체 후보를 포함합니다.
        # 각 후보가 처음 포함되는 최소 영역을 단계로 기록해 중복 OCR 호출을 줄입니다.
        clip = page.rect
        detected = self._ocr_page_detect(page)
        if detected is None:
            self.last_keyword_status = OCR_UNAVAILABLE
            self.last_keyword_reason = self.ocr_unavailable_reason
            return []
        for box, text, score in zip(
            detected["boxes"],
            detected["texts"],
            detected["scores"],
        ):
            if (
                score is None
                or score < OCR_MIN_TEXT_SCORE
                or normalized_keyword not in self._normalized_keyword(text)
            ):
                continue
            rect = self._ocr_box_to_pdf_rect(
                box,
                clip,
                detected["width"],
                detected["height"],
                transform_matrix=detected.get("transform_matrix"),
                source_width=detected.get("source_width"),
                source_height=detected.get("source_height"),
            )
            if not clips[-1].intersects(rect):
                continue
            rect = self._keyword_subrect(rect, text, keyword)
            step = next(
                (
                    index
                    for index, step_clip in enumerate(clips, start=1)
                    if step_clip.contains(rect)
                ),
                5,
            )
            candidates.append(
                {
                    "rect": rect,
                    "score": float(score),
                    "step": step,
                }
            )
        return [
            item["rect"]
            for item in self._merge_ocr_candidates(candidates)
        ]

    def find_keyword_candidates(self, page, origin_rect, keyword):
        """네이티브 우선, OCR 폴백으로 5단계 거리순 후보를 반환합니다."""
        self.last_keyword_status = None
        self.last_keyword_reason = None
        if not keyword:
            return []
        candidates = self._find_native_keyword_candidates(
            page,
            origin_rect,
            keyword,
        )
        if candidates:
            self.last_keyword_status = NATIVE_MATCH
        else:
            candidates = self._find_ocr_keyword_candidates(
                page,
                origin_rect,
                keyword,
            )
            if candidates:
                self.last_keyword_status = OCR_MATCH
            elif self.last_keyword_status != OCR_UNAVAILABLE:
                self.last_keyword_status = OCR_NO_MATCH
        return sorted(
            candidates,
            key=lambda match: (
                self._distance(origin_rect, match),
                match.y0,
                match.x0,
            ),
        )

    def create_mapping(
        self,
        template_page,
        drag_rect,
        keyword="",
        anchor_rect=None,
        anchor_index=0,
    ):
        keyword, keyword_is_valid = validate_anchor_keyword(keyword)
        if not keyword_is_valid:
            raise ValueError(
                f"기준 단어는 최대 {MAX_ANCHOR_KEYWORD_LENGTH}자까지 입력할 수 있습니다."
            )
        candidates = self.find_keyword_candidates(template_page, drag_rect, keyword)
        if anchor_rect is None and candidates:
            anchor_index = min(max(int(anchor_index), 0), len(candidates) - 1)
            anchor_rect = candidates[anchor_index]
        elif anchor_rect is not None and candidates:
            anchor_rect = fitz.Rect(anchor_rect)
            anchor_index = min(
                range(len(candidates)),
                key=lambda index: self._distance(anchor_rect, candidates[index]),
            )
        if keyword and anchor_rect is None:
            logging.warning(f"기준 단어를 찾지 못해 절대좌표를 사용합니다: {keyword}")
        return {
            "rect": fitz.Rect(drag_rect),
            "keyword": keyword,
            "anchor_rect": fitz.Rect(anchor_rect) if anchor_rect else None,
            "tracking_anchor_rect": fitz.Rect(anchor_rect) if anchor_rect else None,
            "offset_x": drag_rect.x0 - anchor_rect.x0 if anchor_rect else 0,
            "offset_y": drag_rect.y0 - anchor_rect.y0 if anchor_rect else 0,
        }

    @staticmethod
    def reset_mapping_tracking(mapping):
        """새 PDF 처리를 시작할 때 추적 기준을 첫 페이지에서 확정한 앵커로 복원합니다."""
        template_anchor = mapping.get("anchor_rect")
        mapping["tracking_anchor_rect"] = (
            fitz.Rect(template_anchor) if template_anchor is not None else None
        )

    @staticmethod
    def _tracking_search_origin(page, drag_rect, tracking_anchor):
        """이전 앵커 좌표를 중심으로 드래그 영역 크기의 탐색 기준 영역을 구성합니다."""
        center = fitz.Point(
            (tracking_anchor.x0 + tracking_anchor.x1) / 2,
            (tracking_anchor.y0 + tracking_anchor.y1) / 2,
        )
        width = max(drag_rect.width, tracking_anchor.width * 2, 5)
        height = max(drag_rect.height, tracking_anchor.height * 2, 5)
        return fitz.Rect(
            center.x - width / 2,
            center.y - height / 2,
            center.x + width / 2,
            center.y + height / 2,
        ) & page.rect

    def adjusted_rect(self, page, mapping):
        drag_rect = mapping["rect"]
        keyword = mapping.get("keyword")
        template_anchor = mapping.get("anchor_rect")
        preprocess_result = self.last_preprocess_result
        if not keyword or template_anchor is None:
            if preprocess_result is not None and preprocess_result.orientation_applied:
                preprocess_result.status = "rejected"
                preprocess_result.reference_validation = "rejected"
                preprocess_result.failure_reason = "anchor_required_for_rotation"
                logging.warning("PAGE_REFERENCE_VALIDATION_REJECTED: reason=anchor_required_for_rotation")
                return fitz.Rect()
            return drag_rect
        tracking_anchor = mapping.get("tracking_anchor_rect") or template_anchor
        search_origin = self._tracking_search_origin(page, drag_rect, tracking_anchor)
        candidates = self.find_keyword_candidates(page, search_origin, keyword)
        if not candidates:
            if preprocess_result is not None and preprocess_result.orientation_applied:
                preprocess_result.status = "rejected"
                preprocess_result.reference_validation = "rejected"
                preprocess_result.failure_reason = "anchor_mismatch"
                logging.warning("PAGE_REFERENCE_VALIDATION_REJECTED: reason=anchor_mismatch")
                return fitz.Rect()
            if self.last_keyword_status == OCR_NO_MATCH:
                logging.warning(
                    "%s: 파일=%s, 페이지=%s, 키워드=%s, 절대좌표 사용",
                    OCR_NO_MATCH,
                    os.path.basename(getattr(page.parent, "name", "") or "(메모리 PDF)"),
                    page.number + 1,
                    keyword,
                )
            return drag_rect
        if self.last_keyword_status == OCR_MATCH:
            logging.info(
                "%s: 파일=%s, 페이지=%s, 키워드=%s",
                OCR_MATCH,
                os.path.basename(getattr(page.parent, "name", "") or "(메모리 PDF)"),
                page.number + 1,
                keyword,
            )
        # 후보 순번은 페이지 크기 변화에 따라 뒤바뀔 수 있으므로 사용하지 않습니다.
        # 이전 페이지에서 확인된 실제 앵커 좌표에 가장 가까운 후보를 자동 추적합니다.
        current_anchor = min(
            candidates,
            key=lambda candidate: self._distance(tracking_anchor, candidate),
        )
        mapping["tracking_anchor_rect"] = fitz.Rect(current_anchor)
        if preprocess_result is not None and preprocess_result.orientation_applied:
            preprocess_result.reference_validation = "accepted"
            logging.info(
                "PAGE_REFERENCE_VALIDATION_ACCEPTED: 파일=%s, 페이지=%s, 키워드=%s",
                os.path.basename(getattr(page.parent, "name", "") or "(메모리 PDF)"),
                page.number + 1,
                keyword,
            )
            template_center = fitz.Point(
                (template_anchor.x0 + template_anchor.x1) / 2,
                (template_anchor.y0 + template_anchor.y1) / 2,
            )
            drag_center = fitz.Point(
                (drag_rect.x0 + drag_rect.x1) / 2,
                (drag_rect.y0 + drag_rect.y1) / 2,
            )
            dx = drag_center.x - template_center.x
            dy = drag_center.y - template_center.y
            angle = preprocess_result.detected_orientation % 360
            rotated_dx, rotated_dy = {
                90: (-dy, dx),
                180: (-dx, -dy),
                270: (dy, -dx),
            }.get(angle, (dx, dy))
            anchor_center = fitz.Point(
                (current_anchor.x0 + current_anchor.x1) / 2,
                (current_anchor.y0 + current_anchor.y1) / 2,
            )
            width, height = drag_rect.width, drag_rect.height
            if angle in (90, 270):
                width, height = height, width
            center = fitz.Point(
                anchor_center.x + rotated_dx,
                anchor_center.y + rotated_dy,
            )
            return fitz.Rect(
                center.x - width / 2,
                center.y - height / 2,
                center.x + width / 2,
                center.y + height / 2,
            ) & page.rect
        return fitz.Rect(
            current_anchor.x0 + mapping["offset_x"],
            current_anchor.y0 + mapping["offset_y"],
            current_anchor.x0 + mapping["offset_x"] + drag_rect.width,
            current_anchor.y0 + mapping["offset_y"] + drag_rect.height,
        ) & page.rect

    @staticmethod
    def _combine_positioned_text(recognized):
        """Group OCR boxes into visual lines and preserve line breaks."""
        if not recognized:
            return ""
        ordered = sorted(
            recognized,
            key=lambda item: (
                (item[0].y0 + item[0].y1) / 2,
                item[0].x0,
            ),
        )
        lines = []
        for rect, text in ordered:
            center_y = (rect.y0 + rect.y1) / 2
            height = max(rect.height, 1)
            if lines:
                previous = lines[-1]
                tolerance = max(previous["height"], height) * 0.6
                if abs(center_y - previous["center_y"]) <= tolerance:
                    previous["items"].append((rect.x0, text))
                    count = len(previous["items"])
                    previous["center_y"] = (
                        previous["center_y"] * (count - 1) + center_y
                    ) / count
                    previous["height"] = max(previous["height"], height)
                    continue
            lines.append(
                {
                    "center_y": center_y,
                    "height": height,
                    "items": [(rect.x0, text)],
                }
            )
        return "\n".join(
            " ".join(
                text
                for _x, text in sorted(line["items"], key=lambda item: item[0])
            )
            for line in lines
        ).strip()

    def _detected_items(self, detected, target_rect=None):
        """Convert one OCR result to scored PDF-coordinate text items."""
        if not detected:
            return []
        source_clip = fitz.Rect(detected["clip"])
        recognized = []
        for box, text, score in zip(
            detected["boxes"],
            detected["texts"],
            detected["scores"],
        ):
            if (
                not text
                or not str(text).strip()
                or score is None
                or score < OCR_MIN_TEXT_SCORE
            ):
                continue
            rect = self._ocr_box_to_pdf_rect(
                box,
                source_clip,
                detected["width"],
                detected["height"],
                transform_matrix=detected.get("transform_matrix"),
                source_width=detected.get("source_width"),
                source_height=detected.get("source_height"),
            )
            if target_rect is not None:
                center = fitz.Point(
                    (rect.x0 + rect.x1) / 2,
                    (rect.y0 + rect.y1) / 2,
                )
                if not target_rect.contains(center):
                    continue
            recognized.append(
                {
                    "rect": rect,
                    "text": str(text).strip(),
                    "score": float(score),
                }
            )
        return recognized

    def _recognized_consensus_text(
        self,
        page_detected,
        region_detected,
        target_rect,
    ):
        """Use region OCR to correct matching page boxes without adding edge noise."""
        page_items = self._detected_items(page_detected, target_rect)
        region_items = self._detected_items(region_detected)
        if not page_items:
            return self._combine_positioned_text(
                [(item["rect"], item["text"]) for item in region_items]
            )

        selected = []
        remaining = list(region_items)
        for page_item in page_items:
            page_rect = page_item["rect"]
            page_center_y = (page_rect.y0 + page_rect.y1) / 2
            horizontal_padding = max(page_rect.height * 0.5, 2)
            matches = []
            for candidate in remaining:
                candidate_rect = candidate["rect"]
                candidate_center_x = (
                    candidate_rect.x0 + candidate_rect.x1
                ) / 2
                candidate_center_y = (
                    candidate_rect.y0 + candidate_rect.y1
                ) / 2
                vertical_tolerance = max(
                    page_rect.height,
                    candidate_rect.height,
                    3,
                ) * 0.6
                if (
                    page_rect.x0 - horizontal_padding
                    <= candidate_center_x
                    <= page_rect.x1 + horizontal_padding
                    and abs(candidate_center_y - page_center_y)
                    <= vertical_tolerance
                ):
                    matches.append(candidate)

            chosen = page_item
            if matches:
                for match in matches:
                    remaining.remove(match)
                match_text = self._combine_positioned_text(
                    [(item["rect"], item["text"]) for item in matches]
                )
                match_score = sum(
                    item["score"] for item in matches
                ) / len(matches)
                page_length = len(self._normalized_keyword(page_item["text"]))
                match_length = len(self._normalized_keyword(match_text))
                if (
                    match_score > page_item["score"]
                    or (
                        match_score >= page_item["score"] - 0.05
                        and match_length > page_length * 1.5
                    )
                ):
                    chosen = {
                        "text": match_text,
                        "score": match_score,
                    }
            selected.append((page_rect, chosen["text"]))
        return self._combine_positioned_text(selected)

    def extract_text(
        self,
        page,
        mapping,
        force_ocr=False,
        dpi=OCR_VALUE_RECOGNITION_DPI,
    ):
        if force_ocr:
            self._ocr_page_detect(page)
        rect = self.adjusted_rect(page, mapping)
        if rect.width < 5 or rect.height < 5:
            logging.warning("추출 영역이 5pt 미만이므로 판독을 생략합니다.")
            return ""
        if not force_ocr:
            try:
                native_text = page.get_text("text", clip=rect).replace("\n", " ").strip()
                if native_text:
                    return native_text
            except Exception as error:
                logging.warning(f"Native 텍스트 추출 실패, OCR로 전환합니다: {error}")
        if not self.ensure_ocr():
            return ""
        try:
            page_detected = self._ocr_page_detect(page)
            region_detected = self._ocr_detect(page, rect, dpi)
            return self._recognized_consensus_text(
                page_detected,
                region_detected,
                rect,
            )
        except Exception as error:
            self._set_ocr_unavailable("OCR_EXECUTION_ERROR", str(error))
            logging.debug("OCR 판독 예외", exc_info=True)
            return ""


TEXT_EXTRACTOR = HybridTextExtractor()


def _make_unique_headers(headers):
    """표시명이 같은 열이 있어도 내부 매핑이 충돌하지 않도록 열 번호를 붙입니다."""
    totals = {}
    for header in headers:
        if header:
            totals[header] = totals.get(header, 0) + 1
    return [
        f"{header} [열 {column_idx}]" if header and totals[header] > 1 else header
        for column_idx, header in enumerate(headers, start=1)
    ]


def _build_merged_write_lookup(ws):
    """병합 영역의 각 좌표를 실제 쓰기 가능한 좌상단 좌표로 연결합니다."""
    return {
        (row, column): (merged_range.min_row, merged_range.min_col)
        for merged_range in ws.merged_cells.ranges
        for row in range(merged_range.min_row, merged_range.max_row + 1)
        for column in range(merged_range.min_col, merged_range.max_col + 1)
    }


def _extract_headers(ws, start_row, end_row):
    merged_values = _build_merged_value_lookup(ws)
    headers = []
    for column in range(1, ws.max_column + 1):
        texts = []
        for row in range(start_row, end_row + 1):
            value = ws.cell(row=row, column=column).value
            if value is None:
                value = merged_values.get((row, column))
            text = str(value).strip() if value is not None else ""
            if text and (not texts or texts[-1] != text):
                texts.append(text)
        headers.append("_".join(texts))
    return _make_unique_headers(headers)


def _rect_to_list(rect):
    if rect is None:
        return None
    rect = fitz.Rect(rect)
    return [rect.x0, rect.y0, rect.x1, rect.y1]


def _mapping_to_json(header, column, mapping):
    anchor = mapping.get("anchor_rect")
    keyword = (mapping.get("keyword") or "").strip() or None
    return {
        "column": column,
        "header": header,
        "keyword": keyword,
        "rect": _rect_to_list(mapping["rect"]),
        "anchor_rect": _rect_to_list(anchor),
        "anchor_offset": (
            [mapping.get("offset_x", 0), mapping.get("offset_y", 0)]
            if anchor is not None else None
        ),
    }


def _mapping_from_json(field):
    keyword = field.get("keyword")
    anchor_rect = field.get("anchor_rect")
    anchor_offset = field.get("anchor_offset")
    anchor_values = (keyword, anchor_rect, anchor_offset)
    if any(value is not None for value in anchor_values) and not all(
        value is not None for value in anchor_values
    ):
        raise ValueError(
            f"'{field.get('header', '이름 없음')}' 필드의 앵커 정보가 일부만 존재합니다."
        )
    rect_values = field.get("rect")
    if not isinstance(rect_values, list) or len(rect_values) != 4:
        raise ValueError(f"'{field.get('header', '이름 없음')}' 필드의 영역 좌표가 올바르지 않습니다.")
    if anchor_offset is not None and (
        not isinstance(anchor_offset, list) or len(anchor_offset) != 2
    ):
        raise ValueError(f"'{field.get('header', '이름 없음')}' 필드의 앵커 간격이 올바르지 않습니다.")
    rect = fitz.Rect(rect_values)
    anchor = fitz.Rect(anchor_rect) if anchor_rect is not None else None
    return {
        "rect": rect,
        "keyword": keyword or "",
        "anchor_rect": anchor,
        "tracking_anchor_rect": fitz.Rect(anchor) if anchor is not None else None,
        "offset_x": anchor_offset[0] if anchor_offset is not None else 0,
        "offset_y": anchor_offset[1] if anchor_offset is not None else 0,
    }


def _native_page_metrics(page):
    """짧은 장비명만 있는 페이지를 문서형으로 오판하지 않도록 지표를 계산합니다."""
    text = page.get_text("text").strip()
    valid_characters = sum(character.isalnum() for character in text)
    blocks = [
        block
        for block in page.get_text("blocks")
        if len(block) > 4 and str(block[4]).strip()
    ]
    return valid_characters, len(blocks)


def select_reference_page(pdf_paths):
    """앞 3페이지에서 네이티브 문서를 우선 추천하고 회전·스캔은 폴백합니다."""
    scanned_fallback = None
    rotated_pages = []
    unreadable_files = []
    for pdf_path in pdf_paths:
        try:
            with fitz.open(pdf_path) as document:
                for page_index in range(
                    min(document.page_count, REFERENCE_SCAN_MAX_PAGES)
                ):
                    page = document[page_index]
                    rotation = int(page.rotation) % 360
                    valid_characters, block_count = _native_page_metrics(page)
                    candidate = {
                        "pdf_path": pdf_path,
                        "page_index": page_index,
                        "source_type": (
                            "native"
                            if (
                                not rotation
                                and valid_characters >= MIN_NATIVE_TEXT_CHARS
                                and block_count >= MIN_NATIVE_TEXT_BLOCKS
                            )
                            else "scanned"
                        ),
                        "text_length": valid_characters,
                        "block_count": block_count,
                        "page_width": page.rect.width,
                        "page_height": page.rect.height,
                        "rotation": rotation,
                    }
                    if candidate["source_type"] == "native":
                        logging.info(
                            "PDF 기준 페이지 선정: 파일=%s, 페이지=%s, "
                            "판독유형=PDF 텍스트",
                            os.path.basename(pdf_path),
                            page_index + 1,
                        )
                        return candidate
                    if scanned_fallback is None:
                        scanned_fallback = candidate
        except Exception as error:
            unreadable_files.append(
                f"{os.path.basename(pdf_path)}: {error}"
            )
            logging.warning(
                "기준 페이지 사전 점검에서 PDF를 읽지 못했습니다: %s, 사유=%s",
                os.path.basename(pdf_path),
                error,
            )
    if scanned_fallback is not None:
        logging.info(
            "PDF 기준 페이지 선정: 파일=%s, 페이지=%s, 판독유형=이미지 OCR",
            os.path.basename(scanned_fallback["pdf_path"]),
            scanned_fallback["page_index"] + 1,
        )
        return scanned_fallback
    rotated_description = ", ".join(
        f"{os.path.basename(item['pdf_path'])} {item['page_index'] + 1}페이지 "
        f"({item['rotation']}도)"
        for item in rotated_pages
    )
    raise ValueError(
        "기준 페이지를 선정할 수 없습니다. "
        "사용 가능한 비회전 PDF 페이지가 없습니다."
        + (f"\n\n{rotated_description}" if rotated_description else "")
        + (
            "\n\n읽기 실패:\n" + "\n".join(unreadable_files)
            if unreadable_files
            else ""
        )
    )

class ReferenceReselectionRequested(RuntimeError):
    """사용자가 OCR 기준 페이지 대신 다른 페이지 탐색을 선택했습니다."""


class ReferencePageSelector:
    """자동 추천값에서 시작해 모든 선택 PDF 페이지를 수동 탐색합니다."""

    def __init__(self, pdf_paths, suggestion, parent_root=None):
        page_counts = []
        valid_paths = []
        for path in pdf_paths:
            with fitz.open(path) as document:
                if document.page_count:
                    valid_paths.append(path)
                    page_counts.append(document.page_count)
        if not valid_paths:
            raise ValueError("기준 페이지로 탐색할 PDF가 없습니다.")
        self.state = ReferenceNavigationState.from_suggestion(
            valid_paths, page_counts, suggestion
        )
        self.result = None
        self.document = None
        self.root = tk.Toplevel(parent_root)
        self.root.title("기준 파일·기준 페이지 선택")
        self.root.transient(parent_root)
        self.root.protocol("WM_DELETE_WINDOW", self._close)
        self.root.geometry("960x780")
        self.root.minsize(720, 560)

        info = tk.Frame(self.root, bg="#E3F2FD", padx=12, pady=10)
        info.pack(fill=tk.X)
        self.page_label = tk.Label(
            info, bg="#E3F2FD", fg="#0D47A1", font=("맑은 고딕", 11, "bold")
        )
        self.page_label.pack(anchor="w")
        self.type_label = tk.Label(
            info, bg="#E3F2FD", fg="#37474F", font=("맑은 고딕", 10)
        )
        self.type_label.pack(anchor="w", pady=(4, 0))

        preview_frame = tk.Frame(self.root, bg="#777777")
        preview_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        self.canvas = tk.Canvas(preview_frame, bg="#777777", highlightthickness=0)
        self.v_scroll = tk.Scrollbar(
            preview_frame, orient=tk.VERTICAL, command=self.canvas.yview
        )
        self.canvas.configure(yscrollcommand=self.v_scroll.set)
        self.v_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self._resize_job = None
        self._rendered_width = 0
        self.canvas.bind("<Configure>", self._on_canvas_configure)
        self.canvas.bind("<MouseWheel>", self._on_mouse_wheel)
        controls = tk.Frame(self.root, padx=10, pady=8)
        controls.pack(fill=tk.X)
        self.previous_file_button = tk.Button(
            controls, text="이전 파일", command=lambda: self._move("previous_file")
        )
        self.previous_page_button = tk.Button(
            controls, text="이전 페이지", command=lambda: self._move("previous_page")
        )
        self.next_page_button = tk.Button(
            controls, text="다음 페이지", command=lambda: self._move("next_page")
        )
        self.next_file_button = tk.Button(
            controls, text="다음 파일", command=lambda: self._move("next_file")
        )
        for button in (
            self.previous_file_button,
            self.previous_page_button,
            self.next_page_button,
            self.next_file_button,
        ):
            button.pack(side=tk.LEFT, padx=3)
        tk.Button(
            controls,
            text="이 페이지를 기준으로 선택",
            command=self._confirm,
            bg="#2E7D32",
            fg="white",
            padx=12,
        ).pack(side=tk.RIGHT, padx=3)
        tk.Button(controls, text="취소", command=self._close).pack(side=tk.RIGHT, padx=3)

        self._render()
        self.root.grab_set()
        try:
            self.root.wait_window()
        finally:
            if self.document is not None:
                self.document.close()
                self.document = None

    def _candidate(self):
        page = self.document[self.state.page_index]
        valid_characters, block_count = _native_page_metrics(page)
        source_type = (
            "native"
            if valid_characters >= MIN_NATIVE_TEXT_CHARS
            and block_count >= MIN_NATIVE_TEXT_BLOCKS
            else "scanned"
        )
        return {
            "pdf_path": self.state.current_path,
            "page_index": self.state.page_index,
            "source_type": source_type,
            "text_length": valid_characters,
            "block_count": block_count,
            "page_width": page.rect.width,
            "page_height": page.rect.height,
            "rotation": int(page.rotation) % 360,
        }

    def _render_preview(self, reset_scroll=False, force=False):
        if self.document is None or not self.root.winfo_exists():
            return
        canvas_width = self.canvas.winfo_width()
        if canvas_width <= 1:
            return
        target_width = max(canvas_width - 2, 1)
        if not force and abs(target_width - self._rendered_width) < 2:
            return

        previous_y = self.canvas.yview()[0] if self.canvas.bbox("all") else 0.0
        page = self.document[self.state.page_index]
        page_width = max(float(page.rect.width), 1.0)
        scale = target_width / page_width
        pix = page.get_pixmap(matrix=fitz.Matrix(scale, scale), alpha=False)
        image = Image.frombytes("RGB", (pix.width, pix.height), pix.samples)
        self.preview_image = ImageTk.PhotoImage(image)
        self.canvas.delete("all")
        self.canvas.create_image(0, 0, anchor=tk.NW, image=self.preview_image)
        self.canvas.configure(scrollregion=(0, 0, pix.width, pix.height))
        self._rendered_width = target_width
        self.canvas.xview_moveto(0)
        self.canvas.yview_moveto(0 if reset_scroll else previous_y)

    def _on_canvas_configure(self, event):
        if event.width <= 1 or abs((event.width - 2) - self._rendered_width) < 2:
            return
        if self._resize_job is not None:
            self.root.after_cancel(self._resize_job)
        self._resize_job = self.root.after(100, self._finish_resize_render)

    def _finish_resize_render(self):
        self._resize_job = None
        self._render_preview(reset_scroll=False)

    def _on_mouse_wheel(self, event):
        if event.delta:
            direction = -1 if event.delta > 0 else 1
            self.canvas.yview_scroll(direction * 3, "units")
        return "break"

    def _render(self):
        if self.document is not None:
            self.document.close()
        self.document = fitz.open(self.state.current_path)
        candidate = self._candidate()
        self.page_label.configure(
            text=(
                f"기준 페이지: {os.path.basename(self.state.current_path)} / "
                f"{self.state.page_index + 1}페이지"
            )
        )
        self.type_label.configure(
            text=(
                "판독 유형: [PDF 텍스트]"
                if candidate["source_type"] == "native"
                else "판독 유형: [OCR 필요]"
            )
        )
        self.previous_file_button.configure(
            state=tk.NORMAL if self.state.can_previous_file else tk.DISABLED
        )
        self.previous_page_button.configure(
            state=tk.NORMAL if self.state.can_previous_page else tk.DISABLED
        )
        self.next_page_button.configure(
            state=tk.NORMAL if self.state.can_next_page else tk.DISABLED
        )
        self.next_file_button.configure(
            state=tk.NORMAL if self.state.can_next_file else tk.DISABLED
        )
        self.root.update_idletasks()
        self._render_preview(reset_scroll=True, force=True)

    def _move(self, method_name):
        getattr(self.state, method_name)()
        self._render()

    def _confirm(self):
        self.result = self._candidate()
        self._close()

    def _close(self):
        try:
            self.root.grab_release()
        except tk.TclError:
            pass
        if self.root.winfo_exists():
            self.root.destroy()

class VisualSelector:
    """
    사용자가 선택한 PDF 대장의 첫 페이지를 GUI에 표시하고
    마우스 드래그를 통해 데이터 추출 영역을 획득하는 마법사 클래스
    """
    def __init__(
        self,
        pdf_path,
        header_name,
        parent_root=None,
        page_index=0,
        source_type="native",
        ocr_confirmation_state=None,
    ):
        self.root = tk.Toplevel(parent_root)
        self.header_name = header_name
        self.source_type = source_type
        self.ocr_confirmation_state = (
            ocr_confirmation_state if ocr_confirmation_state is not None else {}
        )
        self.request_reselection = False
        self.root.title(f"영역 설정 마법사: [{header_name}]")
        self.root.attributes("-topmost", True)

        # 화면의 가로세로 85% 크기로 GUI 창 크기 자동 조절
        sw = self.root.winfo_screenwidth()
        sh = self.root.winfo_screenheight()
        self.root.geometry(f"{int(sw*0.85)}x{int(sh*0.85)}")

        # 상단 조작 가이드 안내 영역
        guide_frame = tk.Frame(self.root, bg="#FFF3E0", pady=15)
        guide_frame.pack(side=tk.TOP, fill=tk.X)

        tk.Label(guide_frame, text=f"항목: [{header_name}] 추출 범위 지정",
                 font=("맑은 고딕", 14, "bold"), fg="#E65100", bg="#FFF3E0").pack()
        tk.Label(
            guide_frame,
            text=(
                f"기준 페이지: {os.path.basename(pdf_path)} / {page_index + 1}페이지 · "
                f"판독 유형: {'PDF 텍스트' if source_type == 'native' else '이미지 OCR'}"
            ),
            font=("맑은 고딕", 9, "bold"),
            fg="#455A64",
            bg="#FFF3E0",
        ).pack(pady=(3, 0))
        tk.Label(guide_frame, text="▶ 방법: 마우스 왼쪽 버튼으로 추출할 데이터가 위치한 '표의 칸'을 드래그하세요.\n"
                                  "▶ 팁: 내용이 길어질 수 있는 항목은 아래쪽 여백을 넉넉하게 잡는 것이 안전합니다.\n"
                                  "▶ 패스: 이 항목을 빈칸으로 남기려면 드래그 없이 우측 상단 [X]를 눌러 창을 닫으세요.",
                 font=("맑은 고딕", 10), bg="#FFF3E0", justify=tk.LEFT).pack()

        # 메인 캔버스 영역 및 스크롤바 바인딩
        self.frame = tk.Frame(self.root)
        self.frame.pack(expand=True, fill=tk.BOTH)

        self.canvas = tk.Canvas(self.frame, bg="white", cursor="cross")
        self.v_scroll = tk.Scrollbar(self.frame, orient=tk.VERTICAL, command=self.canvas.yview)
        self.h_scroll = tk.Scrollbar(self.frame, orient=tk.HORIZONTAL, command=self.canvas.xview)
        self.canvas.configure(yscrollcommand=self.v_scroll.set, xscrollcommand=self.h_scroll.set)

        self.v_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.h_scroll.pack(side=tk.BOTTOM, fill=tk.X)
        self.canvas.pack(side=tk.LEFT, expand=True, fill=tk.BOTH)

        # PDF 파일 최초 로드 및 이미지 변환 (1.8배 해상도 배율 적용)
        self.doc = fitz.open(pdf_path)
        self.page = self.doc[page_index]
        self.zoom = 1.8
        mat = fitz.Matrix(self.zoom, self.zoom)
        pix = self.page.get_pixmap(matrix=mat)

        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        self.tk_img = ImageTk.PhotoImage(img)
        self.canvas.create_image(0, 0, anchor=tk.NW, image=self.tk_img)
        self.canvas.config(scrollregion=(0, 0, pix.width, pix.height))

        self.start_x = self.start_y = 0
        self.rect = None
        self.final_rect = None
        self.anchor_keyword = ""
        self.anchor_candidates = []
        self.anchor_candidate_index = 0
        self.selected_anchor_rect = None
        self.selected_anchor_index = 0
        self.anchor_highlight = None
        self.anchor_search_job = None
        self.focus_animation_job = None
        self.anchor_frame = None

        # 마우스 상호작용 바인딩
        self.canvas.bind("<ButtonPress-1>", self.on_press)
        self.canvas.bind("<B1-Motion>", self.on_drag)
        self.canvas.bind("<ButtonRelease-1>", self.on_release)
        self.canvas.bind("<MouseWheel>", self.on_mouse_wheel)
        self.root.wait_window()
        self.doc.close()
        if self.request_reselection:
            raise ReferenceReselectionRequested()

    def on_press(self, event):
        self.start_x = self.canvas.canvasx(event.x)
        self.start_y = self.canvas.canvasy(event.y)
        if self.rect:
            self.canvas.delete(self.rect)
        self.rect = self.canvas.create_rectangle(self.start_x, self.start_y, self.start_x, self.start_y, outline="#D32F2F", width=3)

    def on_drag(self, event):
        cur_x, cur_y = self.canvas.canvasx(event.x), self.canvas.canvasy(event.y)
        self.canvas.coords(self.rect, self.start_x, self.start_y, cur_x, cur_y)

    def on_release(self, event):
        end_x, end_y = self.canvas.canvasx(event.x), self.canvas.canvasy(event.y)
        x0, y0 = min(self.start_x, end_x) / self.zoom, min(self.start_y, end_y) / self.zoom
        x1, y1 = max(self.start_x, end_x) / self.zoom, max(self.start_y, end_y) / self.zoom
        self.final_rect = fitz.Rect(x0, y0, x1, y1)

        msg_root = tk.Toplevel(self.root)
        msg_root.withdraw()
        msg_root.attributes("-topmost", True)

        confirm = messagebox.askyesno("영역 확정",
            f"항목명: [{self.header_name}]\n"
            f"선택 좌표: {self.final_rect}\n\n"
            "이 영역을 모든 PDF 페이지에 동일하게 적용하시겠습니까?\n"
            "(내용이 칸을 벗어날 것 같으면 '아니오'를 누르고 다시 잡으세요.)", parent=msg_root)

        msg_root.destroy()
        if confirm:
            self._show_anchor_panel()
        else:
            self.final_rect = None
            self.canvas.delete(self.rect)
            self.rect = None

    def _show_anchor_panel(self):
        """캔버스와 분리되지 않는 앵커 검증 패널을 표시합니다."""
        self.canvas.unbind("<ButtonPress-1>")
        self.canvas.unbind("<B1-Motion>")
        self.canvas.unbind("<ButtonRelease-1>")
        self.canvas.configure(cursor="arrow")

        if self.anchor_frame is not None:
            return

        self.anchor_frame = tk.Frame(
            self.root,
            bg="#E8F5E9",
            padx=12,
            pady=10,
            highlightthickness=1,
            highlightbackground="#A5D6A7",
        )
        self.anchor_frame.pack(side=tk.BOTTOM, fill=tk.X, before=self.frame)

        tk.Label(
            self.anchor_frame,
            text="기준 단어(최대 10자)",
            font=("맑은 고딕", 10, "bold"),
            bg="#E8F5E9",
            fg="#1B5E20",
        ).pack(side=tk.LEFT, padx=(0, 8))

        self.anchor_entry = tk.Entry(self.anchor_frame, font=("맑은 고딕", 10), width=24)
        self.anchor_entry.pack(side=tk.LEFT, padx=(0, 10))
        self.anchor_entry.bind("<KeyRelease>", self._schedule_anchor_search)
        self.anchor_entry.bind("<Return>", self._search_anchor_now)

        self.anchor_status = tk.Label(
            self.anchor_frame,
            text="단어를 입력하면 주변 영역을 5단계까지 자동 탐색합니다.",
            font=("맑은 고딕", 9),
            bg="#E8F5E9",
            fg="#455A64",
            anchor="w",
        )
        self.anchor_status.pack(side=tk.LEFT, fill=tk.X, expand=True)

        tk.Button(
            self.anchor_frame,
            text="다음",
            command=self._next_anchor_candidate,
            width=8,
            bg="#1976D2",
            fg="white",
            activebackground="#1565C0",
            activeforeground="white",
        ).pack(side=tk.RIGHT, padx=4)
        tk.Button(
            self.anchor_frame,
            text="확정",
            command=self._confirm_anchor,
            width=8,
            bg="#2E7D32",
            fg="white",
            activebackground="#1B5E20",
            activeforeground="white",
        ).pack(side=tk.RIGHT, padx=4)
        tk.Button(
            self.anchor_frame,
            text="취소",
            command=self._cancel_anchor,
            width=8,
            bg="#757575",
            fg="white",
            activebackground="#616161",
            activeforeground="white",
        ).pack(side=tk.RIGHT, padx=4)

        self.anchor_entry.focus_set()

    def _schedule_anchor_search(self, _event=None):
        if self.anchor_search_job is not None:
            self.root.after_cancel(self.anchor_search_job)
        self.anchor_candidates = []
        self.anchor_candidate_index = 0
        self.selected_anchor_rect = None
        self.selected_anchor_index = 0
        self._clear_anchor_highlight()
        self.anchor_status.configure(text="5단계 탐색 준비 중...", fg="#455A64")
        self.anchor_search_job = self.root.after(250, self._search_anchor_candidates)

    def _search_anchor_now(self, _event=None):
        if self.anchor_search_job is not None:
            self.root.after_cancel(self.anchor_search_job)
            self.anchor_search_job = None
        self._search_anchor_candidates()

    def _search_anchor_candidates(self, _event=None):
        self.anchor_search_job = None
        keyword, keyword_is_valid = validate_anchor_keyword(self.anchor_entry.get())
        self.anchor_candidates = []
        self.anchor_candidate_index = 0
        self.selected_anchor_rect = None
        self.selected_anchor_index = 0
        self._clear_anchor_highlight()

        if not keyword:
            self.anchor_status.configure(
                text="기준 단어를 사용하지 않으려면 [확정]을 선택하십시오.",
                fg="#455A64",
            )
            return
        if not keyword_is_valid:
            self.anchor_status.configure(
                text=(
                    f"기준 단어는 최대 {MAX_ANCHOR_KEYWORD_LENGTH}자입니다. "
                    f"현재 {len(keyword)}자 · 중간 공백 포함"
                ),
                fg="#C62828",
            )
            return

        # 1~5단계 사이에는 대화상자를 표시하지 않고 전체 후보군을 한 번에 완성합니다.
        self.anchor_candidates = TEXT_EXTRACTOR.find_keyword_candidates(
            self.page,
            self.final_rect,
            keyword,
        )
        if not self.anchor_candidates:
            if TEXT_EXTRACTOR.last_keyword_status == OCR_UNAVAILABLE:
                status_text = (
                    "OCR 엔진을 사용할 수 없어 후보를 찾지 못했습니다. "
                    "감사 로그를 확인하십시오."
                )
            else:
                status_text = (
                    "5단계 탐색 완료 · 후보 없음 · "
                    "단어를 수정하거나 [취소]를 선택하십시오."
                )
            self.anchor_status.configure(
                text=status_text,
                fg="#C62828",
            )
            return

        self._show_anchor_candidate(0)

    def _next_anchor_candidate(self):
        if not self.anchor_candidates:
            self._search_anchor_now()
            return
        next_index = (self.anchor_candidate_index + 1) % len(self.anchor_candidates)
        self._show_anchor_candidate(next_index)

    def _show_anchor_candidate(self, index):
        self.anchor_candidate_index = index
        candidate = self.anchor_candidates[index]
        self.selected_anchor_rect = fitz.Rect(candidate)
        self.selected_anchor_index = index
        self._clear_anchor_highlight()

        x0, y0, x1, y1 = (
            candidate.x0 * self.zoom,
            candidate.y0 * self.zoom,
            candidate.x1 * self.zoom,
            candidate.y1 * self.zoom,
        )
        self.anchor_highlight = self.canvas.create_rectangle(
            x0,
            y0,
            x1,
            y1,
            outline="#1976D2",
            width=4,
            tags=("anchor-highlight",),
        )
        self.canvas.tag_raise(self.anchor_highlight)
        self.anchor_status.configure(
            text=(
                f"5단계 탐색 완료 · 후보 {index + 1}/{len(self.anchor_candidates)} · "
                f"거리순 · "
                f"{'OCR' if TEXT_EXTRACTOR.last_keyword_status == OCR_MATCH else 'PDF 텍스트'}"
            ),
            fg="#1B5E20",
        )
        self._focus_canvas_on(candidate)

    def _clear_anchor_highlight(self):
        if self.anchor_highlight is not None:
            self.canvas.delete(self.anchor_highlight)
            self.anchor_highlight = None

    def _focus_canvas_on(self, candidate):
        """선택 후보가 화면 중앙에 오도록 캔버스를 짧게 보간 이동합니다."""
        self.root.update_idletasks()
        content_width = max(float(self.tk_img.width()), 1.0)
        content_height = max(float(self.tk_img.height()), 1.0)
        viewport_width = max(float(self.canvas.winfo_width()), 1.0)
        viewport_height = max(float(self.canvas.winfo_height()), 1.0)
        center_x = (candidate.x0 + candidate.x1) * self.zoom / 2
        center_y = (candidate.y0 + candidate.y1) * self.zoom / 2
        target_x = max(0.0, min(center_x - viewport_width / 2, content_width - viewport_width))
        target_y = max(0.0, min(center_y - viewport_height / 2, content_height - viewport_height))
        target_x_fraction = target_x / content_width
        target_y_fraction = target_y / content_height
        start_x_fraction = self.canvas.xview()[0]
        start_y_fraction = self.canvas.yview()[0]

        if self.focus_animation_job is not None:
            self.root.after_cancel(self.focus_animation_job)
            self.focus_animation_job = None

        def animate(step=1, total_steps=8):
            ratio = step / total_steps
            eased = 1 - (1 - ratio) ** 3
            self.canvas.xview_moveto(
                start_x_fraction + (target_x_fraction - start_x_fraction) * eased
            )
            self.canvas.yview_moveto(
                start_y_fraction + (target_y_fraction - start_y_fraction) * eased
            )
            if step < total_steps:
                self.focus_animation_job = self.root.after(
                    18,
                    lambda: animate(step + 1, total_steps),
                )
            else:
                self.focus_animation_job = None

        animate()

    def _confirm_anchor(self):
        keyword, keyword_is_valid = validate_anchor_keyword(self.anchor_entry.get())
        if not keyword_is_valid:
            if self.anchor_search_job is not None:
                self.root.after_cancel(self.anchor_search_job)
                self.anchor_search_job = None
            self.anchor_status.configure(
                text=(
                    f"기준 단어는 최대 {MAX_ANCHOR_KEYWORD_LENGTH}자입니다. "
                    f"현재 {len(keyword)}자 · 중간 공백 포함"
                ),
                fg="#C62828",
            )
            return
        if self.anchor_search_job is not None:
            self._search_anchor_now()
        if keyword and not self.anchor_candidates:
            self.anchor_status.configure(
                text="후보가 없습니다. 기준 단어를 수정하거나 [취소]를 선택하십시오.",
                fg="#C62828",
            )
            return
        if (
            self.source_type == "scanned"
            and not self.ocr_confirmation_state.get("confirmed")
        ):
            proceed = messagebox.askyesno(
                "OCR 기준 페이지 안내",
                "이 기준 페이지는 이미지 기반(OCR)입니다.\n"
                "문자 인식 특성상 텍스트 기반 문서보다 위치·키워드 인식 정확도가\n"
                "달라질 수 있습니다.\n\n계속 진행하시겠습니까?\n"
                "[아니오]를 누르면 다른 페이지 선택으로 돌아갑니다.",
                parent=self.root,
            )
            if not proceed:
                self.request_reselection = True
                self.root.destroy()
                return
            self.ocr_confirmation_state["confirmed"] = True
        self.anchor_keyword = keyword
        logging.info(
            "좌표 및 기준 단어 확정 -> 항목: %s, Rect: %s, Keyword: %s, "
            "Candidate: %s/%s, SearchStatus: %s",
            self.header_name,
            self.final_rect,
            self.anchor_keyword or "(미사용)",
            self.selected_anchor_index + 1 if self.anchor_candidates else 0,
            len(self.anchor_candidates),
            TEXT_EXTRACTOR.last_keyword_status,
        )
        self.root.destroy()

    def _cancel_anchor(self):
        """기준 단어 입력만 취소하고 확정된 드래그 영역은 절대좌표로 유지합니다."""
        if self.anchor_search_job is not None:
            self.root.after_cancel(self.anchor_search_job)
            self.anchor_search_job = None
        self.anchor_keyword = ""
        self.anchor_candidates = []
        self.selected_anchor_rect = None
        self.selected_anchor_index = 0
        logging.info(
            "기준 단어 입력 취소 -> 절대좌표 사용: %s",
            self.final_rect,
        )
        self.root.destroy()

    def on_mouse_wheel(self, event):
        # Shift 키가 눌려있을 경우 가로 스크롤
        if event.state & 0x0001:
            self.canvas.xview_scroll(int(-1 * (event.delta / 120)), "units")
        else:
            self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

class SheetGroupSelector:
    """PDF 취합 시트·그룹·헤더 범위와 프로파일을 한 화면에서 설정합니다."""

    def __init__(self, template_path, parent_root=None):
        self.template_path = template_path
        self.root = tk.Toplevel(parent_root)
        self.root.title("PDF 취합 시트 및 헤더 설정")
        self.root.configure(bg="#F5F5F5")
        self.root.attributes("-topmost", True)
        self.root.grab_set()
        self.root.geometry("1180x680")
        self.root.minsize(980, 560)

        self.confirmed = False
        self.result = {}
        self.loaded_profile = None
        self.loaded_profile_path = None
        self.workbook = None
        self.rows = {}
        self.current_sheet = None
        self._syncing = False
        self._closed = False
        self._preview_after_id = None
        self._load_result_queue = queue.Queue(maxsize=1)

        info = tk.Frame(
            self.root,
            bg="#E8F5E9",
            highlightthickness=1,
            highlightbackground="#C8E6C9",
        )
        info.pack(fill=tk.X, padx=12, pady=(12, 6))
        tk.Label(
            info,
            text="PDF 취합 시트 및 헤더 설정",
            font=("맑은 고딕", 12, "bold"),
            fg="#2E7D32",
            bg="#E8F5E9",
        ).pack(anchor="w", padx=10, pady=(8, 2))
        tk.Label(
            info,
            text=(
                "취합할 시트와 그룹, 헤더 시작·끝 행을 지정하세요. "
                "같은 그룹의 S·E는 자동으로 동기화됩니다."
            ),
            font=("맑은 고딕", 9),
            fg="#388E3C",
            bg="#E8F5E9",
        ).pack(anchor="w", padx=10, pady=(0, 8))

        self.loading_var = tk.StringVar(value="취합양식을 불러오는 중입니다.")
        self.loading = ttk.Progressbar(self.root, mode="indeterminate")
        self.loading.pack(fill=tk.X, padx=12, pady=(0, 4))
        self.loading.start(12)
        tk.Label(
            self.root,
            textvariable=self.loading_var,
            bg="#F5F5F5",
            fg="#555555",
            anchor="w",
        ).pack(fill=tk.X, padx=14)

        self.paned = ttk.Panedwindow(self.root, orient=tk.HORIZONTAL)
        self.paned.pack(expand=True, fill=tk.BOTH, padx=12, pady=8)
        self.grid_container = tk.Frame(self.paned, bg="white")
        self.preview_container = tk.Frame(self.paned, bg="white")
        self.paned.add(self.grid_container, weight=3)
        self.paned.add(self.preview_container, weight=4)
        self._build_preview()

        buttons = tk.Frame(self.root, bg="#F5F5F5")
        buttons.pack(fill=tk.X, padx=12, pady=(0, 12))
        self.profile_label_var = tk.StringVar(value="적용 프로파일: 없음")
        tk.Label(
            buttons,
            textvariable=self.profile_label_var,
            bg="#F5F5F5",
            fg="#555555",
        ).pack(side=tk.LEFT)
        self.load_button = tk.Button(
            buttons,
            text="프로파일 불러오기",
            command=self.load_profile,
            state=tk.DISABLED,
        )
        self.load_button.pack(side=tk.LEFT, padx=8)
        self.confirm_button = tk.Button(
            buttons,
            text="설정 완료",
            command=self.on_confirm,
            width=12,
            bg="#2E7D32",
            fg="white",
            state=tk.DISABLED,
        )
        self.confirm_button.pack(side=tk.RIGHT, padx=(8, 0))
        tk.Button(
            buttons,
            text="취소",
            command=self.on_cancel,
            width=10,
            bg="#9E9E9E",
            fg="white",
        ).pack(side=tk.RIGHT)

        self.root.protocol("WM_DELETE_WINDOW", self.on_cancel)
        threading.Thread(target=self._load_workbook, daemon=True).start()
        self.root.after(50, self._poll_workbook_load)
        self.root.wait_window()

    def _build_preview(self):
        self.preview_panel = SheetPreviewPanel(
            self.preview_container,
            column_label_mode="number",
        )
        # 기존 GUI 시험과 내부 호출부가 사용하는 공개 속성을 유지합니다.
        self.preview = self.preview_panel.tree
        self.preview_status_var = self.preview_panel.status_var

    def _load_workbook(self):
        try:
            workbook = load_workbook(
                self.template_path,
                data_only=True,
                keep_vba=self.template_path.lower().endswith(".xlsm"),
            )
            result = ("ok", workbook)
        except Exception as error:
            result = ("error", error)
        if self._closed:
            if result[0] == "ok":
                result[1].close()
            return
        self._load_result_queue.put(result)

    def _poll_workbook_load(self):
        if self._closed:
            return
        try:
            result = self._load_result_queue.get_nowait()
        except queue.Empty:
            self.root.after(50, self._poll_workbook_load)
            return
        self._finish_workbook_load(result)

    def _finish_workbook_load(self, result):
        if self._closed:
            if result[0] == "ok":
                result[1].close()
            return
        self.loading.stop()
        self.loading.pack_forget()
        if result[0] == "error":
            error = result[1]
            logging.error("PDF 설정용 워크북 로딩 실패: %s", error, exc_info=True)
            self.loading_var.set("취합양식을 불러오지 못했습니다.")
            self.preview_status_var.set("미리보기를 사용할 수 없습니다.")
            messagebox.showerror(
                PROGRAM_NAME,
                f"취합양식을 불러오지 못했습니다.\n\n파일: {self.template_path}\n사유: {error}",
                parent=self.root,
            )
            return
        self.workbook = result[1]
        self.loading_var.set("시트와 헤더 범위를 설정해 주세요.")
        self._build_grid(self.workbook.sheetnames)
        self.load_button.config(state=tk.NORMAL)
        self.confirm_button.config(state=tk.NORMAL)
        if self.workbook.sheetnames:
            self.select_sheet(self.workbook.sheetnames[0])

    def _build_grid(self, sheet_names):
        column_specs = (
            (0, 42, 0),
            (1, 120, 3),
            (2, 90, 2),
            (3, 55, 1),
            (4, 55, 1),
        )
        table_frame = tk.Frame(self.grid_container, bg="white")
        table_frame.pack(expand=True, fill=tk.BOTH, padx=6, pady=(8, 8))
        table_frame.columnconfigure(0, weight=1)
        table_frame.rowconfigure(1, weight=1)

        header = tk.Frame(table_frame, bg="#E0E0E0")
        header.grid(row=0, column=0, sticky="ew")
        for column, minimum, weight in column_specs:
            header.columnconfigure(column, minsize=minimum, weight=weight)
        for column, text in enumerate(("선택", "시트명", "그룹", "S행", "E행")):
            tk.Label(
                header,
                text=text,
                bg="#E0E0E0",
                font=("맑은 고딕", 9, "bold"),
                anchor="center",
                padx=3,
                pady=4,
            ).grid(row=0, column=column, sticky="nsew")

        canvas = tk.Canvas(table_frame, bg="white", highlightthickness=0)
        scrollbar = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=canvas.yview)
        body = tk.Frame(canvas, bg="white")
        body.bind(
            "<Configure>",
            lambda _event: canvas.configure(scrollregion=canvas.bbox("all")),
        )
        body_window = canvas.create_window((0, 0), window=body, anchor="nw")
        canvas.bind(
            "<Configure>",
            lambda event: canvas.itemconfigure(body_window, width=event.width),
        )
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.grid(row=1, column=0, sticky="nsew")
        scrollbar.grid(row=0, column=1, rowspan=2, sticky="ns")

        for sheet_name in sheet_names:
            row_frame = tk.Frame(body, bg="white")
            row_frame.pack(fill=tk.X, pady=2)
            for column, minimum, weight in column_specs:
                row_frame.columnconfigure(column, minsize=minimum, weight=weight)
            selected = tk.BooleanVar(value=True)
            group = tk.StringVar(value="그룹 1")
            start = tk.StringVar(value="1")
            end = tk.StringVar(value="")
            check = tk.Checkbutton(
                row_frame,
                variable=selected,
                bg="white",
                command=lambda name=sheet_name: self._toggle_sheet(name),
            )
            check.grid(row=0, column=0, sticky="nsew", padx=2)
            label = tk.Label(
                row_frame,
                text=sheet_name,
                anchor="center",
                bg="white",
                cursor="hand2",
                padx=4,
            )
            label.grid(row=0, column=1, sticky="nsew", padx=2)
            label.bind("<Button-1>", lambda _event, name=sheet_name: self.select_sheet(name))
            group_menu = tk.OptionMenu(
                row_frame,
                group,
                *GROUP_OPTIONS,
                command=lambda _value, name=sheet_name: self._group_changed(name),
            )
            group_menu.config(anchor="center")
            group_menu.grid(row=0, column=2, sticky="ew", padx=2)
            start_entry = tk.Entry(row_frame, textvariable=start, justify="center")
            start_entry.grid(row=0, column=3, sticky="ew", padx=2)
            end_entry = tk.Entry(row_frame, textvariable=end, justify="center")
            end_entry.grid(row=0, column=4, sticky="ew", padx=2)
            for widget in (start_entry, end_entry):
                widget.bind("<FocusIn>", lambda _event, name=sheet_name: self.select_sheet(name))
            start.trace_add(
                "write",
                lambda *_args, name=sheet_name: self._range_changed(name),
            )
            end.trace_add(
                "write",
                lambda *_args, name=sheet_name: self._range_changed(name),
            )
            self.rows[sheet_name] = {
                "selected": selected,
                "group": group,
                "start": start,
                "end": end,
                "widgets": (group_menu, start_entry, end_entry),
            }

    def _toggle_sheet(self, sheet_name):
        row = self.rows[sheet_name]
        state = tk.NORMAL if row["selected"].get() else tk.DISABLED
        for widget in row["widgets"]:
            widget.config(state=state)
        self.select_sheet(sheet_name)

    def _group_changed(self, sheet_name):
        if self._syncing:
            return
        row = self.rows[sheet_name]
        group = row["group"].get()
        if group != "개별":
            existing = next(
                (
                    other
                    for name, other in self.rows.items()
                    if name != sheet_name
                    and other["selected"].get()
                    and other["group"].get() == group
                ),
                None,
            )
            if existing is not None:
                self._syncing = True
                row["start"].set(existing["start"].get())
                row["end"].set(existing["end"].get())
                self._syncing = False
        self.select_sheet(sheet_name)
        self._schedule_preview()

    def _range_changed(self, sheet_name):
        if self._syncing or sheet_name not in self.rows:
            return
        row = self.rows[sheet_name]
        group = row["group"].get()
        if group != "개별" and row["selected"].get():
            self._syncing = True
            for name, other in self.rows.items():
                if (
                    name != sheet_name
                    and other["selected"].get()
                    and other["group"].get() == group
                ):
                    other["start"].set(row["start"].get())
                    other["end"].set(row["end"].get())
            self._syncing = False
        if self.current_sheet == sheet_name:
            self._schedule_preview()

    def select_sheet(self, sheet_name):
        self.current_sheet = sheet_name
        self._schedule_preview()

    def _schedule_preview(self):
        if self._preview_after_id is not None:
            try:
                self.root.after_cancel(self._preview_after_id)
            except tk.TclError:
                pass
        self._preview_after_id = self.root.after(
            HEADER_PREVIEW_DEBOUNCE_MS,
            self._refresh_preview,
        )

    @staticmethod
    def _parse_positive_int(value):
        stripped = (value or "").strip()
        if not stripped:
            return None
        number = int(stripped)
        if number < 1 or number > EXCEL_MAX_ROW:
            raise ValueError
        return number

    def _refresh_preview(self):
        self._preview_after_id = None
        if self.workbook is None or self.current_sheet not in self.rows:
            return
        row_config = self.rows[self.current_sheet]
        try:
            start = self._parse_positive_int(row_config["start"].get())
            end = self._parse_positive_int(row_config["end"].get())
        except ValueError:
            self.preview_status_var.set("S·E에는 1 이상의 정수를 입력해 주세요.")
            return
        if start is None:
            self.preview_status_var.set("S행을 입력하면 미리보기가 표시됩니다.")
            return
        if end is not None and end < start:
            self.preview_status_var.set("E행은 S행보다 크거나 같아야 합니다.")
            return

        self.preview_panel.refresh(
            self.workbook[self.current_sheet],
            self.current_sheet,
            start,
            end,
        )

    def load_profile(self):
        initial_dir = profile_directory(PDF_PROFILE_TYPE)
        path = filedialog.askopenfilename(
            parent=self.root,
            title="PDF 매핑 프로파일 불러오기",
            initialdir=initial_dir if os.path.isdir(initial_dir) else application_dir(),
            filetypes=[("PDF 매핑 프로파일", "*.json")],
        )
        if not path:
            return
        try:
            profile, legacy_pdf = read_profile(
                path,
                PDF_PROFILE_TYPE,
                allow_legacy_pdf=True,
            )
            fatal, minor = self._validate_profile_structure(profile)
            if fatal:
                raise ValueError("\n".join(f"- {reason}" for reason in fatal))
            if legacy_pdf:
                minor.insert(
                    0,
                    "이 파일은 profile_type이 없는 이전 PDF 프로파일입니다. "
                    "적용은 가능하지만 다시 저장하면 새 형식으로 변환됩니다.",
                )
            if minor and not messagebox.askyesno(
                PROGRAM_NAME,
                "프로파일과 현재 파일에 차이가 있습니다.\n\n"
                + "\n".join(f"- {reason}" for reason in minor)
                + "\n\n저장된 설정을 계속 적용하시겠습니까?",
                parent=self.root,
            ):
                logging.info("프로파일 경미한 불일치 확인 후 사용자가 적용을 취소했습니다: %s", path)
                return
            self._apply_profile(profile)
            self.loaded_profile = profile
            self.loaded_profile_path = path
            self.profile_label_var.set(
                f"적용 프로파일: {profile.get('metadata', {}).get('profile_name') or os.path.basename(path)}"
            )
            logging.info(
                "PDF 매핑 프로파일 불러오기: %s, 경미한 불일치=%s",
                path,
                minor,
            )
        except Exception as error:
            logging.error("PDF 매핑 프로파일 불러오기 실패: %s", error, exc_info=True)
            messagebox.showerror(
                PROGRAM_NAME,
                f"프로파일을 적용할 수 없습니다.\n\n사유:\n{error}",
                parent=self.root,
            )

    def _validate_profile_structure(self, profile):
        fatal = []
        minor = []
        if not isinstance(profile, dict):
            return ["프로파일 최상위 구조가 객체가 아닙니다."], minor
        if profile.get("profile_type") not in (None, PDF_PROFILE_TYPE):
            fatal.append(
                f"PDF 매핑 프로파일이 아닙니다: {profile.get('profile_type')}"
            )
        if profile.get("schema_version") != PROFILE_SCHEMA_VERSION:
            fatal.append(
                f"지원하지 않는 스키마 버전입니다: {profile.get('schema_version')}"
            )
        mapping_sets = profile.get("mapping_sets")
        if not isinstance(mapping_sets, list) or not mapping_sets:
            fatal.append("mapping_sets가 없거나 비어 있습니다.")
            return fatal, minor
        known_sheets = set(self.workbook.sheetnames)
        seen_profile_sheets = set()
        for index, mapping_set in enumerate(mapping_sets, start=1):
            if not isinstance(mapping_set, dict):
                fatal.append(f"{index}번째 매핑 구조가 객체가 아닙니다.")
                continue
            group = mapping_set.get("group")
            sheets = mapping_set.get("sheets")
            if group not in GROUP_OPTIONS:
                fatal.append(f"{index}번째 매핑의 그룹 값이 올바르지 않습니다: {group}")
            if not isinstance(sheets, list) or not sheets:
                fatal.append(f"{index}번째 매핑에 시트가 없습니다.")
                continue
            if group == "개별" and len(sheets) != 1:
                fatal.append(f"개별 매핑 {index}에는 시트 하나만 있어야 합니다.")
            for sheet_name in sheets:
                if sheet_name not in known_sheets:
                    fatal.append(f"프로파일의 시트 '{sheet_name}'이 취합양식에 없습니다.")
                if sheet_name in seen_profile_sheets:
                    fatal.append(f"시트 '{sheet_name}'이 둘 이상의 매핑에 중복 지정되었습니다.")
                seen_profile_sheets.add(sheet_name)
            try:
                start = int(mapping_set.get("header_start"))
                end = int(mapping_set.get("header_end"))
                if start < 1 or end < start or end > EXCEL_MAX_ROW:
                    raise ValueError
            except (TypeError, ValueError):
                fatal.append(f"{index}번째 매핑의 S·E 범위가 올바르지 않습니다.")
            fields = mapping_set.get("fields")
            if not isinstance(fields, list) or not fields:
                fatal.append(f"{index}번째 매핑에 PDF 필드가 없습니다.")
                continue
            seen_columns = set()
            for field in fields:
                column = field.get("column")
                if not isinstance(column, int) or column < 1:
                    fatal.append(f"{index}번째 매핑에 유효하지 않은 열 번호가 있습니다.")
                elif column in seen_columns:
                    fatal.append(f"{index}번째 매핑에서 {column}열이 중복 지정되었습니다.")
                seen_columns.add(column)
                try:
                    _mapping_from_json(field)
                except ValueError as error:
                    fatal.append(str(error))
            pdf_info = mapping_set.get("pdf")
            if not isinstance(pdf_info, dict) or not all(
                key in pdf_info for key in ("page_width", "page_height", "rotation")
            ):
                fatal.append(f"{index}번째 매핑에 기준 PDF 페이지 정보가 없습니다.")
        metadata = profile.get("metadata", {})
        template_name = metadata.get("template_file_name")
        if template_name and template_name != os.path.basename(self.template_path):
            minor.append(
                f"기준 취합양식은 '{template_name}'이고 현재 파일은 "
                f"'{os.path.basename(self.template_path)}'입니다."
            )
        if metadata.get("app_version") not in (None, VERSION):
            minor.append(
                f"프로파일 생성 버전은 {metadata.get('app_version')}이고 "
                f"현재 버전은 {VERSION}입니다."
            )
        expected_field_count = sum(
            len(item.get("fields", []))
            for item in mapping_sets
            if isinstance(item, dict)
        )
        if metadata.get("field_count") not in (None, expected_field_count):
            minor.append(
                f"메타데이터의 필드 수({metadata.get('field_count')})와 "
                f"실제 필드 수({expected_field_count})가 다릅니다."
            )
        return fatal, minor

    def _apply_profile(self, profile):
        self._syncing = True
        for row in self.rows.values():
            row["selected"].set(False)
        for mapping_set in profile["mapping_sets"]:
            for sheet_name in mapping_set["sheets"]:
                row = self.rows[sheet_name]
                row["selected"].set(True)
                row["group"].set(mapping_set["group"])
                row["start"].set(str(mapping_set["header_start"]))
                row["end"].set(str(mapping_set["header_end"]))
                for widget in row["widgets"]:
                    widget.config(state=tk.NORMAL)
        self._syncing = False
        first_sheet = profile["mapping_sets"][0]["sheets"][0]
        self.select_sheet(first_sheet)

    def on_confirm(self):
        result = {}
        errors = []
        for sheet_name, row in self.rows.items():
            if not row["selected"].get():
                continue
            try:
                start = self._parse_positive_int(row["start"].get())
                end = self._parse_positive_int(row["end"].get())
                if start is None or end is None:
                    raise ValueError("S행과 E행은 모두 필수 입력입니다.")
                if end < start:
                    raise ValueError("E행은 S행보다 크거나 같아야 합니다.")
                result[sheet_name] = {
                    "group": row["group"].get(),
                    "S": start,
                    "E": end,
                }
            except (TypeError, ValueError) as error:
                reason = str(error) or "1~1,048,576 범위의 정수를 입력해 주세요."
                errors.append(f"{sheet_name}: {reason}")
        if not result:
            errors.append("최소 하나의 시트를 선택해 주세요.")
        if errors:
            messagebox.showerror(
                PROGRAM_NAME,
                "설정을 완료할 수 없습니다.\n\n" + "\n".join(f"- {item}" for item in errors),
                parent=self.root,
            )
            return
        self.result = result
        self.confirmed = True
        self._close()

    def on_cancel(self):
        self.confirmed = False
        self._close()

    def _close(self):
        self._closed = True
        if self._preview_after_id is not None:
            try:
                self.root.after_cancel(self._preview_after_id)
            except tk.TclError:
                pass
        if self.workbook is not None:
            try:
                self.workbook.close()
            except Exception as error:
                logging.warning("PDF 설정용 워크북 닫기 실패: %s", error)
            self.workbook = None
        try:
            pending = self._load_result_queue.get_nowait()
            if pending[0] == "ok":
                pending[1].close()
        except queue.Empty:
            pass
        self.root.destroy()

def _build_mapping_sets(sheet_settings):
    """그룹 시트는 공유 단위로, 개별 시트는 독립 단위로 묶습니다."""
    groups = {}
    individual = []
    for sheet_name, settings in sheet_settings.items():
        group = settings["group"]
        item = {
            "group": group,
            "sheets": [sheet_name],
            "S": settings["S"],
            "E": settings["E"],
        }
        if group == "개별":
            individual.append(item)
        else:
            existing = groups.setdefault(group, item)
            if existing is not item:
                existing["sheets"].append(sheet_name)
    return list(groups.values()) + individual


def _find_loaded_mapping_set(profile, current_set):
    if not profile:
        return None
    current_sheets = set(current_set["sheets"])
    return next(
        (
            item for item in profile.get("mapping_sets", [])
            if item.get("group") == current_set["group"]
            and set(item.get("sheets", [])) == current_sheets
        ),
        None,
    )


def _load_profile_mapping(mapping_set, headers, reference, parent_root):
    """현재 헤더와 PDF 규격을 검증한 뒤 JSON 매핑을 실행 객체로 복원합니다."""
    pdf_path = reference["pdf_path"]
    page_index = reference["page_index"]
    fatal = []
    minor = []
    fields = {field.get("column"): field for field in mapping_set.get("fields", [])}
    mapping = {}
    for column, header in enumerate(headers, start=1):
        if not header:
            continue
        field = fields.get(column)
        if field is None:
            fatal.append(f"{column}열 '{header}'의 매핑이 프로파일에 없습니다.")
            continue
        if field.get("header") != header:
            minor.append(
                f"{column}열 헤더가 '{field.get('header')}'에서 '{header}'(으)로 변경되었습니다."
            )
        try:
            mapping[header] = _mapping_from_json(field)
        except ValueError as error:
            fatal.append(str(error))

    with fitz.open(pdf_path) as document:
        if not document.page_count:
            fatal.append(f"PDF에 페이지가 없습니다: {os.path.basename(pdf_path)}")
        else:
            page = document[page_index]
            pdf_info = mapping_set.get("pdf") or {}
            page_info_valid = True
            try:
                page_mismatch = (
                    abs(float(pdf_info["page_width"]) - page.rect.width) > 1
                    or abs(float(pdf_info["page_height"]) - page.rect.height) > 1
                    or int(pdf_info["rotation"]) != int(page.rotation)
                )
            except (KeyError, TypeError, ValueError):
                page_mismatch = True
                page_info_valid = False
                fatal.append("프로파일에 기준 PDF 페이지 정보가 없습니다.")
            if page_mismatch and page_info_valid:
                fatal.append(
                    "현재 PDF의 페이지 크기 또는 회전 상태가 프로파일 기준과 다릅니다. "
                    f"현재={page.rect.width:.2f}×{page.rect.height:.2f}, {page.rotation}도"
                )
            for header, item_mapping in mapping.items():
                if not page.rect.contains(item_mapping["rect"]):
                    fatal.append(f"'{header}' 필드의 영역 좌표가 현재 PDF 범위를 벗어납니다.")
                anchor = item_mapping.get("anchor_rect")
                if anchor is not None and not page.rect.contains(anchor):
                    fatal.append(f"'{header}' 필드의 앵커 좌표가 현재 PDF 범위를 벗어납니다.")

    reference_name = (mapping_set.get("pdf") or {}).get("reference_pdf_name")
    if reference_name and reference_name != os.path.basename(pdf_path):
        minor.append(
            f"기준 PDF는 '{reference_name}'이고 현재 파일은 '{os.path.basename(pdf_path)}'입니다."
        )
    if fatal:
        raise ValueError("\n".join(f"- {reason}" for reason in fatal))
    if minor:
        logging.warning("PDF 프로파일 경미한 불일치: %s", minor)
        if not messagebox.askyesno(
            PROGRAM_NAME,
            "프로파일과 현재 파일에 차이가 있습니다.\n\n"
            + "\n".join(f"- {reason}" for reason in minor)
            + "\n\n저장된 매핑을 계속 적용하시겠습니까?",
            parent=parent_root,
        ):
            raise RuntimeError("사용자가 프로파일 적용을 취소했습니다.")
    return mapping


def _create_visual_mapping(reference, headers, parent_root):
    mapping = {}
    ocr_confirmation_state = {}
    pdf_path = reference["pdf_path"]
    page_index = reference["page_index"]
    with fitz.open(pdf_path) as document:
        if not document.page_count:
            raise ValueError(f"PDF에 페이지가 없습니다: {os.path.basename(pdf_path)}")
        page = document[page_index]
        for header in headers:
            if not header:
                continue
            selector = VisualSelector(
                pdf_path,
                header,
                parent_root=parent_root,
                page_index=page_index,
                source_type=reference["source_type"],
                ocr_confirmation_state=ocr_confirmation_state,
            )
            if selector.final_rect:
                mapping[header] = TEXT_EXTRACTOR.create_mapping(
                    page,
                    selector.final_rect,
                    selector.anchor_keyword,
                    selector.selected_anchor_rect,
                    selector.selected_anchor_index,
                )
    return mapping


def _profile_document(template_path, configured_sets):
    now = datetime.datetime.now(
        datetime.timezone(datetime.timedelta(hours=9))
    ).isoformat()
    serialized_sets = []
    field_count = 0
    for configured in configured_sets:
        spec = configured["spec"]
        fields = [
            _mapping_to_json(header, column, configured["mapping"][header])
            for column, header in enumerate(configured["headers"], start=1)
            if header and header in configured["mapping"]
        ]
        field_count += len(fields)
        reference = configured.get("reference") or select_reference_page(
            [
                pdf_path
                for sheet_name in spec["sheets"]
                for pdf_path in configured["pdfs"][sheet_name]
            ]
        )
        reference_pdf = reference["pdf_path"]
        pdf_info = {
            "reference_pdf_name": os.path.basename(reference_pdf),
            "reference_page_index": reference["page_index"],
            "page_width": reference["page_width"],
            "page_height": reference["page_height"],
            "rotation": reference["rotation"],
        }
        serialized_sets.append(
            {
                "group": spec["group"],
                "sheets": spec["sheets"],
                "header_start": spec["S"],
                "header_end": spec["E"],
                "pdf": pdf_info,
                "fields": fields,
            }
        )
    first = configured_sets[0]
    first_reference = first.get("reference") or select_reference_page(
        [
            pdf_path
            for sheet_name in first["spec"]["sheets"]
            for pdf_path in first["pdfs"][sheet_name]
        ]
    )
    first_pdf = first_reference["pdf_path"]
    return {
        "schema_version": PROFILE_SCHEMA_VERSION,
        "profile_type": PDF_PROFILE_TYPE,
        "metadata": {
            "profile_name": "",
            "created_at": now,
            "updated_at": now,
            "app_version": VERSION,
            "template_file_name": os.path.basename(template_path),
            "reference_pdf_name": os.path.basename(first_pdf),
            "field_count": field_count,
        },
        "mapping_sets": serialized_sets,
    }


def _offer_profile_save(template_path, configured_sets, parent_root):
    if not messagebox.askyesno(
        PROGRAM_NAME,
        "현재 PDF 영역·앵커·시트 설정을 매핑 프로파일로 저장하시겠습니까?",
        parent=parent_root,
    ):
        return
    profile = _profile_document(template_path, configured_sets)
    default_dir = profile_directory(PDF_PROFILE_TYPE)
    try:
        initial_dir = prepare_profile_directory(PDF_PROFILE_TYPE)
    except OSError as error:
        logging.warning("기본 프로파일 폴더 준비 실패: %s", error)
        messagebox.showwarning(
            PROGRAM_NAME,
            "기본 프로파일 폴더에 저장할 수 없습니다.\n\n"
            f"경로: {default_dir}\n사유: {error}\n\n다른 저장 위치를 선택해 주세요.",
            parent=parent_root,
        )
        initial_dir = os.path.dirname(template_path)
    path = filedialog.asksaveasfilename(
        parent=parent_root,
        title="PDF 매핑 프로파일 저장",
        initialdir=initial_dir,
        defaultextension=".json",
        filetypes=[("DnS Auto 프로파일", "*.json")],
    )
    if not path:
        return
    profile["metadata"]["profile_name"] = os.path.splitext(os.path.basename(path))[0]
    try:
        write_profile(profile, path)
        logging.info(
            "PDF 매핑 프로파일 저장: %s, 시트=%s, 필드=%s",
            path,
            sum(len(item["sheets"]) for item in profile["mapping_sets"]),
            profile["metadata"]["field_count"],
        )
        messagebox.showinfo(
            PROGRAM_NAME,
            f"매핑 프로파일을 저장했습니다.\n\n{path}",
            parent=parent_root,
        )
    except Exception as error:
        logging.error("PDF 매핑 프로파일 저장 실패: %s", error, exc_info=True)
        if not messagebox.askyesno(
            PROGRAM_NAME,
            "프로파일을 저장하지 못했습니다.\n\n"
            f"경로: {path}\n사유: {error}\n\n다른 위치에 저장하시겠습니까?",
            parent=parent_root,
        ):
            return
        retry_path = filedialog.asksaveasfilename(
            parent=parent_root,
            title="PDF 매핑 프로파일을 다른 위치에 저장",
            initialdir=os.path.dirname(template_path),
            initialfile=os.path.basename(path),
            defaultextension=".json",
            filetypes=[("DnS Auto 프로파일", "*.json")],
        )
        if not retry_path:
            return
        profile["metadata"]["profile_name"] = os.path.splitext(
            os.path.basename(retry_path)
        )[0]
        try:
            write_profile(profile, retry_path)
            logging.info("PDF 매핑 프로파일 대체 경로 저장: %s", retry_path)
            messagebox.showinfo(
                PROGRAM_NAME,
                f"매핑 프로파일을 저장했습니다.\n\n{retry_path}",
                parent=parent_root,
            )
        except Exception as retry_error:
            logging.error(
                "PDF 매핑 프로파일 대체 경로 저장 실패: %s",
                retry_error,
                exc_info=True,
            )
            messagebox.showerror(
                PROGRAM_NAME,
                "프로파일을 저장하지 못했습니다.\n\n"
                f"경로: {retry_path}\n사유: {retry_error}",
                parent=parent_root,
            )


def _collect_pdf_rows(
    pdf_paths,
    headers,
    mapping,
    force_ocr,
    failed_files,
    processing_summary=None,
    progress=None,
    overall_state=None,
    cancellation=None,
    sheet_name="",
):
    rows = []
    summary = processing_summary if processing_summary is not None else {}
    summary.setdefault("processed_pages", 0)
    summary.setdefault("empty_pages", [])
    summary.setdefault("rotated_pages", [])
    summary.setdefault("ocr_unavailable_pages", [])
    summary.setdefault("deskew_pages", [])
    summary.setdefault("orientation_pages", [])
    summary.setdefault("preprocess_rejected_pages", [])
    for pdf_path in pdf_paths:
        filename = os.path.basename(pdf_path)
        try:
            with fitz.open(pdf_path) as document:
                for item_mapping in mapping.values():
                    TEXT_EXTRACTOR.reset_mapping_tracking(item_mapping)
                pdf_has_text = False
                for page_index, page in enumerate(document):
                    TEXT_EXTRACTOR.last_preprocess_result = None
                    if cancellation is not None and cancellation.should_cancel():
                        break
                    if progress is not None and overall_state is not None:
                        try:
                            has_native_text = bool(
                                page.get_text("text").strip()
                            )
                        except Exception:
                            has_native_text = False
                        work_type = (
                            "ocr" if force_ocr or int(page.rotation) % 360 or not has_native_text else
                            "native_text"
                        )
                        activity = (
                            "PDF 텍스트 추출 중" if work_type == "native_text" else
                            "OCR로 텍스트 추출 중"
                        )
                        overall_state["current"] += 1
                        progress.begin_unit(
                            filename,
                            overall_state["current"],
                            (
                                f"{document.page_count}페이지 중 "
                                f"{page_index + 1}페이지 | {activity}"
                            ),
                            work_type=work_type,
                            ocr_weight=1 + len(mapping),
                            sheet_name=sheet_name,
                        )
                    values = {}
                    page_has_text = False
                    for header in headers:
                        if not header:
                            continue
                        text = (
                            TEXT_EXTRACTOR.extract_text(
                                page,
                                mapping[header],
                                force_ocr=force_ocr or bool(int(page.rotation) % 360),
                            )
                            if header in mapping else ""
                        )
                        values[header] = text
                        if text:
                            page_has_text = True
                            pdf_has_text = True
                    preprocess_result = TEXT_EXTRACTOR.last_preprocess_result
                    if preprocess_result is not None:
                        if preprocess_result.orientation_applied:
                            summary["orientation_pages"].append(
                                {
                                    "file": filename,
                                    "page": page_index + 1,
                                    "detected": preprocess_result.detected_orientation,
                                    "correction": preprocess_result.orientation_correction,
                                    "confidence": round(preprocess_result.orientation_confidence, 4),
                                    "margin": round(preprocess_result.orientation_margin, 4),
                                }
                            )
                        if preprocess_result.deskew_applied:
                            summary["deskew_pages"].append(
                                {
                                    "file": filename,
                                    "page": page_index + 1,
                                    "angle": round(preprocess_result.skew_angle, 4),
                                    "confidence": round(preprocess_result.skew_confidence, 4),
                                    "line_count": preprocess_result.valid_line_count,
                                    "angle_dispersion": round(preprocess_result.angle_dispersion, 4),
                                }
                            )
                        if preprocess_result.status == "rejected":
                            summary["preprocess_rejected_pages"].append(
                                {
                                    "file": filename,
                                    "page": page_index + 1,
                                    "reason": preprocess_result.failure_reason,
                                }
                            )
                    if page_has_text:
                        rows.append(values)
                        summary["processed_pages"] += 1
                    else:
                        empty_page = {
                            "file": filename,
                            "page": page_index + 1,
                        }
                        summary["empty_pages"].append(empty_page)
                        if TEXT_EXTRACTOR.ocr_unavailable_reason:
                            summary["ocr_unavailable_pages"].append(
                                {
                                    **empty_page,
                                    "reason": TEXT_EXTRACTOR.ocr_unavailable_reason,
                                }
                            )
                        logging.warning(
                            "[%s] %s페이지에서 텍스트가 추출되지 않았습니다.",
                            filename,
                            page_index + 1,
                        )
                    if progress is not None and overall_state is not None:
                        progress.complete_unit(overall_state["current"], work_type=work_type, ocr_weight=1 + len(mapping), ocr_initialization_seconds=getattr(TEXT_EXTRACTOR, "last_ocr_initialization_seconds", 0.0))
                        TEXT_EXTRACTOR.last_ocr_initialization_seconds = 0.0
                if not pdf_has_text and filename not in failed_files:
                    failed_files.append(filename)
        except Exception as error:
            logging.error("PDF 파일 로드 및 파싱 실패: %s, 사유: %s", filename, error)
            if filename not in failed_files:
                failed_files.append(filename)
        finally:
            TEXT_EXTRACTOR.release_pdf_cache(pdf_path)
    return rows


def _processing_summary_text(summary, failed_files):
    lines = [
        f"정상 처리: {summary.get('processed_pages', 0)}페이지",
        f"빈 결과: {len(summary.get('empty_pages', []))}페이지",
        f"회전 보정: {len(summary.get('orientation_pages', []))}페이지",
        f"기울기 보정: {len(summary.get('deskew_pages', []))}페이지",
        f"전처리 폐기: {len(summary.get('preprocess_rejected_pages', []))}페이지",
        (
            "OCR 사용 불가로 미처리: "
            f"{len(summary.get('ocr_unavailable_pages', []))}페이지"
        ),
    ]
    if failed_files:
        lines.append(f"제외 파일: {len(failed_files)}개")
    orientation_pages = summary.get("orientation_pages", [])
    if orientation_pages:
        lines.extend(("", "회전 보정 내역:"))
        for item in orientation_pages[:10]:
            lines.append(
                f"- {item['file']} / {item['page']}페이지 / "
                f"{item['detected']}도 감지 → {item['correction']}도 보정"
            )
    deskew_pages = summary.get("deskew_pages", [])
    if deskew_pages:
        lines.extend(("", "기울기 보정 내역:"))
        for item in deskew_pages[:10]:
            lines.append(
                f"- {item['file']} / {item['page']}페이지 / {item['angle']:+.2f}도"
            )
    rotated_pages = summary.get("rotated_pages", [])
    if rotated_pages:
        lines.append("")
        lines.append("회전 제외 내역:")
        for item in rotated_pages[:10]:
            lines.append(
                f"- {item['file']} / {item['page']}페이지 / {item['rotation']}도"
            )
        if len(rotated_pages) > 10:
            lines.append(
                f"- 외 {len(rotated_pages) - 10}페이지: 감사 로그 확인"
            )
    return "\n".join(lines)


def run_application(parent_root, force_ocr=False):
    """PDF 영역 취합을 그리드 설정과 선택적 매핑 프로파일로 실행합니다."""
    logging.info("=== %s PDF Drag 취합 가동 시작 ===", PROGRAM_NAME)
    wb_data = None
    wb_write = None
    progress = None
    try:
        TEXT_EXTRACTOR.reset_work_cache()
        if force_ocr and RapidOCR is None:
            messagebox.showerror(
                PROGRAM_NAME,
                "강제 OCR을 사용할 수 없습니다.\n"
                "rapidocr-onnxruntime 및 onnxruntime 설치 상태를 확인해 주세요.",
                parent=parent_root,
            )
            return
        template_path = filedialog.askopenfilename(
            parent=parent_root,
            title="1. [필수] 데이터 취합용 엑셀 양식을 선택하세요",
            filetypes=[("Excel Files", "*.xlsx *.xlsm *.xls"), ("All Files", "*.*")],
        )
        if not template_path:
            return
        if template_path.lower().endswith(".xls"):
            raise ValueError(
                "구버전 Excel (.xls) 형식은 직접 수정할 수 없습니다.\n"
                "xlsx 또는 xlsm 형식으로 저장한 뒤 사용해 주세요."
            )

        selector = SheetGroupSelector(template_path, parent_root=parent_root)
        if not selector.confirmed:
            logging.info("사용자가 PDF 시트 및 헤더 설정을 취소했습니다.")
            return

        keep_vba = template_path.lower().endswith(".xlsm")
        wb_data = load_workbook(template_path, data_only=True, keep_vba=keep_vba)
        configured_sets = []
        for spec in _build_mapping_sets(selector.result):
            master_sheet = spec["sheets"][0]
            headers = _extract_headers(wb_data[master_sheet], spec["S"], spec["E"])
            if not any(headers):
                raise ValueError(
                    f"[{master_sheet}] {spec['S']}~{spec['E']}행에서 "
                    "유효한 헤더를 찾지 못했습니다."
                )

            pdfs_by_sheet = {}
            for sheet_name in spec["sheets"]:
                selected_pdfs = filedialog.askopenfilenames(
                    parent=parent_root,
                    title=f"[{sheet_name}] 취합할 PDF 파일 선택",
                    filetypes=[("PDF Files", "*.pdf")],
                )
                if not selected_pdfs:
                    logging.info("[%s] PDF 선택을 취소하여 작업을 중단합니다.", sheet_name)
                    return
                pdfs_by_sheet[sheet_name] = list(selected_pdfs)

            all_pdf_paths = [
                pdf_path
                for sheet_name in spec["sheets"]
                for pdf_path in pdfs_by_sheet[sheet_name]
            ]
            reference = select_reference_page(all_pdf_paths)
            loaded_set = _find_loaded_mapping_set(selector.loaded_profile, spec)
            if loaded_set is not None:
                try:
                    mapping = _load_profile_mapping(
                        loaded_set,
                        headers,
                        reference,
                        parent_root,
                    )
                except RuntimeError:
                    return
            else:
                if selector.loaded_profile is not None:
                    raise ValueError(
                        f"현재 설정({spec['group']}: {', '.join(spec['sheets'])})과 "
                        "일치하는 매핑이 프로파일에 없습니다."
                    )
                while True:
                    reference_selector = ReferencePageSelector(
                        all_pdf_paths, reference, parent_root=parent_root
                    )
                    if reference_selector.result is None:
                        logging.info("사용자가 기준 파일·페이지 선택을 취소했습니다.")
                        return
                    reference = reference_selector.result
                    try:
                        mapping = _create_visual_mapping(
                            reference, headers, parent_root
                        )
                    except ReferenceReselectionRequested:
                        logging.info(
                            "사용자가 OCR 기준 페이지 대신 다른 페이지 선택을 요청했습니다."
                        )
                        continue
                    break
            if not mapping:
                raise ValueError(f"[{master_sheet}] 저장된 PDF 영역 매핑이 없습니다.")
            configured_sets.append(
                {
                    "spec": spec,
                    "headers": headers,
                    "mapping": mapping,
                    "pdfs": pdfs_by_sheet,
                    "reference": reference,
                }
            )

        _offer_profile_save(template_path, configured_sets, parent_root)

        total_pages = 0
        planned_work = []
        for configured in configured_sets:
            ocr_weight = 1 + len(configured["mapping"])
            for sheet_name in configured["spec"]["sheets"]:
                for pdf_path in configured["pdfs"][sheet_name]:
                    try:
                        with fitz.open(pdf_path) as document:
                            total_pages += document.page_count
                            for page in document:
                                try:
                                    has_text = bool(page.get_text("text").strip())
                                except Exception:
                                    has_text = False
                                planned_work.append(("ocr" if force_ocr or int(page.rotation) % 360 or not has_text else "native_text", ocr_weight))
                    except Exception:
                        pass
        from processing_cancellation import ProcessingCancellation

        cancellation = ProcessingCancellation()
        progress = ProcessingProgressDialog(
            parent_root, "PDF 취합 진행 상황", total_pages, "페이지",
            planned_work=planned_work,
            cancellation=cancellation,
        )
        overall_state = {"current": 0}

        wb_write = load_workbook(template_path, keep_vba=keep_vba)
        failed_files = []
        processing_summary = {
            "processed_pages": 0,
            "empty_pages": [],
            "rotated_pages": [],
            "ocr_unavailable_pages": [],
            "deskew_pages": [],
            "orientation_pages": [],
            "preprocess_rejected_pages": [],
        }
        any_data_extracted = False
        for configured in configured_sets:
            spec = configured["spec"]
            headers = configured["headers"]
            mapping = configured["mapping"]
            logging.info(
                "PDF 매핑 처리 시작: 그룹=%s, 시트=%s, S=%s, E=%s",
                spec["group"],
                spec["sheets"],
                spec["S"],
                spec["E"],
            )
            for sheet_name in spec["sheets"]:
                sheet_rows = _collect_pdf_rows(
                    configured["pdfs"][sheet_name],
                    headers,
                    mapping,
                    force_ocr,
                    failed_files,
                    processing_summary,
                    progress,
                    overall_state,
                    cancellation,
                    sheet_name,
                )
                if cancellation.should_cancel():
                    break
                if not sheet_rows:
                    continue
                any_data_extracted = True
                ws_write = wb_write[sheet_name]
                merged_writes = _build_merged_write_lookup(ws_write)
                for row_offset, values in enumerate(sheet_rows, start=1):
                    output_row = spec["E"] + row_offset
                    for column, header in enumerate(headers, start=1):
                        if not header:
                            continue
                        target = merged_writes.get((output_row, column))
                        if target is not None and target != (output_row, column):
                            continue
                        ws_write.cell(
                            row=output_row,
                            column=column,
                            value=values.get(header, ""),
                        )
            if cancellation.should_cancel():
                break

        if cancellation.should_cancel():
            progress.close()
            progress = None
            messagebox.showinfo(
                "취소됨",
                "전체 작업이 취소되었습니다.\n이번 작업의 결과 파일은 생성되지 않았습니다.",
                parent=parent_root,
            )
            return False

        if not any_data_extracted:
            messagebox.showwarning(
                PROGRAM_NAME,
                "추출된 데이터가 존재하지 않습니다.\n\n"
                + _processing_summary_text(processing_summary, failed_files),
                parent=parent_root,
            )
            return

        extension = ".xlsm" if keep_vba else ".xlsx"
        output_path = os.path.join(
            os.path.dirname(template_path),
            f"DnS_Auto_Drag취합_{datetime.datetime.now().strftime('%H%M%S_%f')}{extension}",
        )
        if not progress.enter_save_phase():
            progress.close()
            progress = None
            messagebox.showinfo(
                "취소됨",
                "전체 작업이 취소되었습니다.\n이번 작업의 결과 파일은 생성되지 않았습니다.",
                parent=parent_root,
            )
            return False
        wb_write.save(output_path)
        progress.close()
        progress = None
        summary_text = _processing_summary_text(
            processing_summary,
            failed_files,
        )
        logging.info("PDF 취합 완료: %s\n%s", output_path, summary_text)
        if (
            failed_files
            or processing_summary["empty_pages"]
            or processing_summary["rotated_pages"]
            or processing_summary["ocr_unavailable_pages"]
        ):
            failed_section = (
                "\n\n제외 파일:\n" + "\n".join(failed_files)
                if failed_files
                else ""
            )
            messagebox.showwarning(
                PROGRAM_NAME,
                "데이터 취합을 완료했으나 확인이 필요한 항목이 있습니다.\n\n"
                + summary_text
                + failed_section
                + "\n\n전체 세부 사유는 감사 로그를 확인해 주세요.",
                parent=parent_root,
            )
        else:
            messagebox.showinfo(
                PROGRAM_NAME,
                "시트별 데이터 취합을 완료했습니다.\n\n"
                + summary_text
                + f"\n\n저장: {output_path}",
                parent=parent_root,
            )
        os.startfile(os.path.dirname(output_path))
        return True
    except Exception as error:
        logging.error("PDF 취합 치명적 오류: %s", error, exc_info=True)
        messagebox.showerror(
            PROGRAM_NAME,
            f"오류가 발생했습니다.\n\n{error}",
            parent=parent_root,
        )
    finally:
        if progress is not None:
            progress.close()
        for workbook in (wb_data, wb_write):
            if workbook is not None:
                try:
                    workbook.close()
                except Exception as close_error:
                    logging.warning("PDF 취합 워크북 닫기 실패: %s", close_error)
