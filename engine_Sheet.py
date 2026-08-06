"""여러 Excel 파일의 시트 데이터를 기준 양식으로 취합하는 기능."""

import datetime
import logging
import os
import queue
import re
import threading
import time
import tkinter as tk
import zipfile
from tkinter import filedialog, messagebox, ttk

from openpyxl import load_workbook
from utils_progress import ProcessingProgressDialog

from config import PROGRAM_NAME, VERSION
from utils_profiles import (
    PROFILE_SCHEMA_VERSION,
    SHEET_PROFILE_TYPE,
    application_dir,
    prepare_profile_directory,
    profile_directory,
    read_profile,
    write_profile,
)
from utils_sheet_preview import (
    EXCEL_MAX_ROW,
    HEADER_PREVIEW_DEBOUNCE_MS,
    SheetPreviewPanel,
)


def _last_data_row(ws, minimum_row=1):
    """서식만 존재하는 고스트 행을 제외한 마지막 실제 데이터 행을 찾습니다."""
    for row_idx in range(ws.max_row, minimum_row - 1, -1):
        if any(
            ws.cell(row=row_idx, column=col_idx).value not in (None, "")
            for col_idx in range(1, ws.max_column + 1)
        ):
            return row_idx
    return minimum_row


class MultiSheetSelector:
    """엑셀 취합 시트별 옵션과 기준양식 미리보기를 한 화면에서 설정합니다."""

    MODE_OPTIONS = [
        "1. 데이터 누적 적재 방식",
        "2. 데이터 지정 입력 방식 (교차 검증 지원)",
    ]
    KEY_OPTIONS = ["선택 없음"] + [chr(index) for index in range(65, 91)]
    COLUMN_SPECS = (
        (0, 50, 0),
        (1, 145, 2),
        (2, 80, 0),
        (3, 80, 0),
        (4, 230, 3),
        (5, 100, 1),
        (6, 120, 1),
    )

    def __init__(self, parent, template_path, sheet_names):
        self.template_path = template_path
        self.selected_sheets = {}
        self.sheet_vars = {}
        self.workbook = None
        self.current_sheet = None
        self._closed = False
        self._preview_after_id = None
        self._load_result_queue = queue.Queue(maxsize=1)
        self._load_started = time.perf_counter()
        self.preview_load_seconds = None

        self.top = tk.Toplevel(parent)
        self.top.title("다중 시트 선택 및 취합 설정")
        self.top.configure(bg="#F5F5F5")
        self.top.attributes("-topmost", True)
        self.top.grab_set()
        screen_width = self.top.winfo_screenwidth()
        screen_height = self.top.winfo_screenheight()
        width = max(980, min(1500, int(screen_width * 0.92)))
        height = max(620, min(820, int(screen_height * 0.84)))
        x = max(0, (screen_width - width) // 2)
        y = max(0, (screen_height - height) // 2)
        self.top.geometry(f"{width}x{height}+{x}+{y}")
        self.top.minsize(min(980, screen_width), min(620, screen_height))

        info = tk.Frame(
            self.top,
            bg="#E8F5E9",
            highlightthickness=1,
            highlightbackground="#C8E6C9",
        )
        info.pack(fill=tk.X, padx=12, pady=(12, 6))
        tk.Label(
            info,
            text="엑셀 취합 시트 및 헤더 설정",
            font=("맑은 고딕", 12, "bold"),
            fg="#2E7D32",
            bg="#E8F5E9",
        ).pack(anchor="w", padx=10, pady=(8, 2))
        tk.Label(
            info,
            text=(
                "취합할 시트와 헤더 범위, 취합 방식을 지정하세요. "
                "선택한 시트의 E행은 직접 입력해야 합니다."
            ),
            font=("맑은 고딕", 9),
            fg="#388E3C",
            bg="#E8F5E9",
        ).pack(anchor="w", padx=10, pady=(0, 8))

        self.loading_var = tk.StringVar(value="기준양식 미리보기를 불러오는 중입니다.")
        self.loading = ttk.Progressbar(self.top, mode="indeterminate")
        self.loading.pack(fill=tk.X, padx=12, pady=(0, 4))
        self.loading.start(12)
        tk.Label(
            self.top,
            textvariable=self.loading_var,
            bg="#F5F5F5",
            fg="#555555",
            anchor="w",
        ).pack(fill=tk.X, padx=14)

        profile_bar = tk.Frame(self.top, bg="#F5F5F5")
        profile_bar.pack(fill=tk.X, padx=12, pady=(4, 0))
        self.profile_label_var = tk.StringVar(value="적용 프로파일: 없음")
        tk.Label(
            profile_bar,
            textvariable=self.profile_label_var,
            bg="#F5F5F5",
            fg="#455A64",
            anchor="w",
        ).pack(side=tk.LEFT, fill=tk.X, expand=True)
        self.load_profile_button = tk.Button(
            profile_bar,
            text="프로파일 불러오기",
            command=self.load_profile,
            state=tk.DISABLED,
        )
        self.load_profile_button.pack(side=tk.RIGHT, padx=(6, 0))

        self.paned = ttk.Panedwindow(self.top, orient=tk.HORIZONTAL)
        self.paned.pack(expand=True, fill=tk.BOTH, padx=12, pady=8)
        self.grid_container = tk.Frame(self.paned, bg="white")
        self.preview_container = tk.Frame(self.paned, bg="white")
        self.paned.add(self.grid_container, weight=5)
        self.paned.add(self.preview_container, weight=4)
        self._build_grid(sheet_names)
        self.preview_panel = SheetPreviewPanel(
            self.preview_container,
            column_label_mode="letter",
        )
        self.preview = self.preview_panel.tree
        self.preview_status_var = self.preview_panel.status_var

        button_frame = tk.Frame(self.top, bg="#F5F5F5")
        button_frame.pack(fill=tk.X, padx=12, pady=(0, 12))
        tk.Label(
            button_frame,
            text="선택한 시트는 S·E행을 모두 입력해야 합니다.",
            bg="#F5F5F5",
            fg="#555555",
        ).pack(side=tk.LEFT)
        tk.Button(
            button_frame,
            text="완료",
            command=self.on_ok,
            width=12,
            bg="#2E7D32",
            fg="white",
            font=("맑은 고딕", 10, "bold"),
        ).pack(side=tk.RIGHT, padx=(8, 0))
        tk.Button(
            button_frame,
            text="취소",
            command=self.on_cancel,
            width=10,
            bg="#9E9E9E",
            fg="white",
            font=("맑은 고딕", 10),
        ).pack(side=tk.RIGHT)

        self.top.protocol("WM_DELETE_WINDOW", self.on_cancel)
        if sheet_names:
            self.select_sheet(sheet_names[0])
        threading.Thread(target=self._load_workbook, daemon=True).start()
        self.top.after(50, self._poll_workbook_load)
        self.top.wait_window()

    @staticmethod
    def _configure_columns(frame):
        for column, minimum, weight in MultiSheetSelector.COLUMN_SPECS:
            frame.columnconfigure(column, minsize=minimum, weight=weight)

    def _build_grid(self, sheet_names):
        tk.Label(
            self.grid_container,
            text="시트별 취합 설정",
            font=("맑은 고딕", 10, "bold"),
            bg="white",
        ).pack(anchor="w", padx=8, pady=(8, 4))
        host = tk.Frame(self.grid_container, bg="white")
        host.pack(expand=True, fill=tk.BOTH, padx=8, pady=(0, 8))
        host.columnconfigure(0, weight=1)
        host.rowconfigure(1, weight=1)

        table_width = sum(spec[1] for spec in self.COLUMN_SPECS)
        header_canvas = tk.Canvas(
            host,
            bg="#E0E0E0",
            height=36,
            highlightthickness=0,
        )
        body_canvas = tk.Canvas(host, bg="white", highlightthickness=0)
        y_scroll = ttk.Scrollbar(host, orient=tk.VERTICAL, command=body_canvas.yview)

        def xview(*args):
            header_canvas.xview(*args)
            body_canvas.xview(*args)

        x_scroll = ttk.Scrollbar(host, orient=tk.HORIZONTAL, command=xview)
        body_canvas.configure(
            yscrollcommand=y_scroll.set,
            xscrollcommand=x_scroll.set,
        )
        header_canvas.grid(row=0, column=0, sticky="ew")
        body_canvas.grid(row=1, column=0, sticky="nsew")
        y_scroll.grid(row=0, column=1, rowspan=2, sticky="ns")
        x_scroll.grid(row=2, column=0, sticky="ew")

        header_frame = tk.Frame(header_canvas, bg="#E0E0E0")
        self._configure_columns(header_frame)
        header_window = header_canvas.create_window(
            (0, 0),
            window=header_frame,
            anchor="nw",
            width=table_width,
        )
        headers = (
            "선택",
            "시트명",
            "헤더 S행",
            "헤더 E행",
            "취합 방식",
            "기준 열",
            "빈칸만 채우기",
        )
        for column, text in enumerate(headers):
            tk.Label(
                header_frame,
                text=text,
                bg="#E0E0E0",
                font=("맑은 고딕", 9, "bold"),
                anchor="center",
                padx=3,
                pady=7,
            ).grid(row=0, column=column, sticky="nsew")

        body = tk.Frame(body_canvas, bg="white")
        body_window = body_canvas.create_window(
            (0, 0),
            window=body,
            anchor="nw",
            width=table_width,
        )

        def update_scrollregion(_event=None):
            body_canvas.configure(scrollregion=body_canvas.bbox("all"))

        body.bind("<Configure>", update_scrollregion)

        def resize_table(event):
            content_width = max(table_width, event.width)
            header_canvas.itemconfigure(header_window, width=content_width)
            body_canvas.itemconfigure(body_window, width=content_width)
            header_canvas.configure(scrollregion=(0, 0, content_width, 36))
            update_scrollregion()

        body_canvas.bind("<Configure>", resize_table)

        for sheet_name in sheet_names:
            row_frame = tk.Frame(body, bg="white")
            row_frame.pack(fill=tk.X, pady=2)
            self._configure_columns(row_frame)
            selected = tk.BooleanVar(value=False)
            start = tk.StringVar(value="1")
            end = tk.StringVar(value="")
            mode = tk.StringVar(value=self.MODE_OPTIONS[0])
            key = tk.StringVar(value=self.KEY_OPTIONS[0])
            protect = tk.BooleanVar(value=True)

            check = tk.Checkbutton(
                row_frame,
                variable=selected,
                bg="white",
                command=lambda name=sheet_name: self._selection_changed(name),
            )
            check.grid(row=0, column=0, sticky="nsew", padx=2)
            label = tk.Label(
                row_frame,
                text=sheet_name,
                font=("맑은 고딕", 9),
                anchor="center",
                bg="white",
                cursor="hand2",
                padx=4,
            )
            label.grid(row=0, column=1, sticky="nsew", padx=2)
            label.bind(
                "<Button-1>",
                lambda _event, name=sheet_name: self.select_sheet(name),
            )
            start_entry = tk.Entry(
                row_frame,
                textvariable=start,
                width=6,
                justify="center",
                state=tk.DISABLED,
                font=("맑은 고딕", 9),
            )
            start_entry.grid(row=0, column=2, sticky="ew", padx=2)
            end_entry = tk.Entry(
                row_frame,
                textvariable=end,
                width=6,
                justify="center",
                state=tk.DISABLED,
                font=("맑은 고딕", 9),
            )
            end_entry.grid(row=0, column=3, sticky="ew", padx=2)
            mode_menu = tk.OptionMenu(
                row_frame,
                mode,
                *self.MODE_OPTIONS,
                command=lambda _value, name=sheet_name: self._mode_changed(name),
            )
            mode_menu.config(font=("맑은 고딕", 8), state=tk.DISABLED, anchor="center")
            mode_menu.grid(row=0, column=4, sticky="ew", padx=2)
            key_menu = tk.OptionMenu(
                row_frame,
                key,
                *self.KEY_OPTIONS,
                command=lambda _value, name=sheet_name: self.select_sheet(name),
            )
            key_menu.config(font=("맑은 고딕", 8), state=tk.DISABLED, anchor="center")
            key_menu.grid(row=0, column=5, sticky="ew", padx=2)
            protect_check = tk.Checkbutton(
                row_frame,
                variable=protect,
                bg="white",
                state=tk.DISABLED,
                command=lambda name=sheet_name: self.select_sheet(name),
            )
            protect_check.grid(row=0, column=6, sticky="nsew", padx=2)

            for widget in (start_entry, end_entry):
                widget.bind(
                    "<FocusIn>",
                    lambda _event, name=sheet_name: self.select_sheet(name),
                )
            start.trace_add(
                "write",
                lambda *_args, name=sheet_name: self._range_changed(name),
            )
            end.trace_add(
                "write",
                lambda *_args, name=sheet_name: self._range_changed(name),
            )
            self.sheet_vars[sheet_name] = {
                "check": selected,
                "s": start,
                "e": end,
                "mode": mode,
                "key": key,
                "protect": protect,
                "widgets": (
                    start_entry,
                    end_entry,
                    mode_menu,
                    key_menu,
                    protect_check,
                ),
                "label": label,
            }

    def _selection_changed(self, sheet_name):
        row = self.sheet_vars[sheet_name]
        enabled = row["check"].get()
        standard_state = tk.NORMAL if enabled else tk.DISABLED
        start_entry, end_entry, mode_menu, key_menu, protect_check = row["widgets"]
        for widget in (start_entry, end_entry, mode_menu, protect_check):
            widget.config(state=standard_state)
        key_menu.config(
            state=(
                tk.NORMAL
                if enabled and "2." in row["mode"].get()
                else tk.DISABLED
            )
        )
        self.select_sheet(sheet_name)

    def _mode_changed(self, sheet_name):
        row = self.sheet_vars[sheet_name]
        key_menu = row["widgets"][3]
        key_menu.config(
            state=(
                tk.NORMAL
                if row["check"].get() and "2." in row["mode"].get()
                else tk.DISABLED
            )
        )
        self.select_sheet(sheet_name)

    def _range_changed(self, sheet_name):
        if self.current_sheet == sheet_name:
            self._schedule_preview()

    def select_sheet(self, sheet_name):
        if sheet_name not in self.sheet_vars:
            return
        self.current_sheet = sheet_name
        for name, row in self.sheet_vars.items():
            row["label"].config(
                bg="#E8F5E9" if name == sheet_name else "white",
            )
        self._schedule_preview()

    def _schedule_preview(self):
        if self._closed:
            return
        if self._preview_after_id is not None:
            try:
                self.top.after_cancel(self._preview_after_id)
            except tk.TclError:
                pass
        self._preview_after_id = self.top.after(
            HEADER_PREVIEW_DEBOUNCE_MS,
            self._refresh_preview,
        )

    @staticmethod
    def _parse_positive_int(value):
        stripped = (value or "").strip()
        if not stripped:
            return None
        number = int(stripped)
        if number < 1 or number > EXCEL_MAX_ROW:
            raise ValueError
        return number

    def _refresh_preview(self):
        self._preview_after_id = None
        if self.workbook is None or self.current_sheet not in self.sheet_vars:
            return
        row = self.sheet_vars[self.current_sheet]
        try:
            start = self._parse_positive_int(row["s"].get())
            end = self._parse_positive_int(row["e"].get())
        except ValueError:
            self.preview_status_var.set("S·E에는 1 이상의 정수를 입력해 주세요.")
            return
        if start is None:
            self.preview_status_var.set("S행을 입력하면 미리보기가 표시됩니다.")
            return
        if end is not None and end < start:
            self.preview_status_var.set("E행은 S행보다 크거나 같아야 합니다.")
            return
        self.preview_panel.refresh(
            self.workbook[self.current_sheet],
            self.current_sheet,
            start,
            end,
        )

    def _load_workbook(self):
        try:
            workbook = load_workbook(
                self.template_path,
                data_only=True,
                keep_vba=self.template_path.lower().endswith(".xlsm"),
            )
            result = ("ok", workbook)
        except Exception as error:
            result = ("error", error)
        if self._closed:
            if result[0] == "ok":
                result[1].close()
            return
        self._load_result_queue.put(result)

    def _poll_workbook_load(self):
        if self._closed:
            return
        try:
            result = self._load_result_queue.get_nowait()
        except queue.Empty:
            self.top.after(50, self._poll_workbook_load)
            return
        self._finish_workbook_load(result)

    def _finish_workbook_load(self, result):
        if self._closed:
            if result[0] == "ok":
                result[1].close()
            return
        self.loading.stop()
        self.loading.pack_forget()
        self.preview_load_seconds = round(
            time.perf_counter() - self._load_started,
            3,
        )
        if result[0] == "error":
            error = result[1]
            logging.error("엑셀 설정용 미리보기 로딩 실패: %s", error)
            self.loading_var.set("미리보기를 불러오지 못했지만 취합 설정은 계속할 수 있습니다.")
            self.preview_status_var.set(f"미리보기 사용 불가: {error}")
            self.load_profile_button.config(state=tk.NORMAL)
            messagebox.showwarning(
                PROGRAM_NAME,
                "기준양식 미리보기를 불러오지 못했습니다.\n"
                "시트 설정과 취합은 계속할 수 있습니다.\n\n"
                f"사유: {error}",
                parent=self.top,
            )
            return
        self.workbook = result[1]
        self.loading_var.set(
            f"기준양식 미리보기 준비 완료 ({self.preview_load_seconds:.3f}초)"
        )
        self.load_profile_button.config(state=tk.NORMAL)
        self._schedule_preview()

    def _collect_settings(self, show_errors=True):
        selected_sheets = {}
        for sheet_name, row in self.sheet_vars.items():
            if not row["check"].get():
                continue
            try:
                start = self._parse_positive_int(row["s"].get())
                end = self._parse_positive_int(row["e"].get())
                if start is None or end is None or end < start:
                    raise ValueError
            except (TypeError, ValueError):
                if show_errors:
                    messagebox.showerror(
                        "입력 오류",
                        f"[{sheet_name}] 시트의 S행과 E행을 모두 입력해 주세요.\n"
                        "(S >= 1, E >= S)",
                        parent=self.top,
                    )
                self.select_sheet(sheet_name)
                return None
            mode = 1 if "1." in row["mode"].get() else 2
            key = row["key"].get()
            selected_sheets[sheet_name] = {
                "S": start,
                "E": end,
                "mode": mode,
                "key_col": "" if key == "선택 없음" else key,
                "protect": row["protect"].get(),
            }

        if not selected_sheets:
            if show_errors:
                messagebox.showwarning(
                    "선택 오류",
                    "최소 하나의 시트를 선택해 주세요.",
                    parent=self.top,
                )
            return None
        return selected_sheets

    def _profile_document(self, selected_sheets):
        now = datetime.datetime.now(
            datetime.timezone(datetime.timedelta(hours=9))
        ).isoformat()
        return {
            "schema_version": PROFILE_SCHEMA_VERSION,
            "profile_type": SHEET_PROFILE_TYPE,
            "metadata": {
                "profile_name": "",
                "created_at": now,
                "updated_at": now,
                "app_version": VERSION,
                "template_file_name": os.path.basename(self.template_path),
                "sheet_count": len(selected_sheets),
            },
            "sheet_configs": [
                {
                    "sheet_name": sheet_name,
                    "header_start": config["S"],
                    "header_end": config["E"],
                    "mode": config["mode"],
                    "key_col": config["key_col"],
                    "protect": bool(config["protect"]),
                }
                for sheet_name, config in selected_sheets.items()
            ],
        }

    def _validate_profile(self, profile):
        fatal = []
        minor = []
        if profile.get("profile_type") != SHEET_PROFILE_TYPE:
            fatal.append(
                f"Sheet 설정 프로파일이 아닙니다: {profile.get('profile_type')}"
            )
        if profile.get("schema_version") != PROFILE_SCHEMA_VERSION:
            fatal.append(
                f"지원하지 않는 스키마 버전입니다: {profile.get('schema_version')}"
            )
        configs = profile.get("sheet_configs")
        if not isinstance(configs, list) or not configs:
            fatal.append("sheet_configs가 없거나 비어 있습니다.")
            return fatal, minor

        known_sheets = set(self.sheet_vars)
        seen_sheets = set()
        for index, config in enumerate(configs, start=1):
            if not isinstance(config, dict):
                fatal.append(f"{index}번째 시트 설정이 객체가 아닙니다.")
                continue
            sheet_name = config.get("sheet_name")
            if sheet_name not in known_sheets:
                fatal.append(f"프로파일의 시트 '{sheet_name}'이 기준양식에 없습니다.")
            if sheet_name in seen_sheets:
                fatal.append(f"시트 '{sheet_name}'이 프로파일에 중복 지정되었습니다.")
            seen_sheets.add(sheet_name)
            try:
                start = int(config.get("header_start"))
                end = int(config.get("header_end"))
                if start < 1 or end < start or end > EXCEL_MAX_ROW:
                    raise ValueError
            except (TypeError, ValueError):
                fatal.append(f"[{sheet_name}] S·E 범위가 올바르지 않습니다.")
            mode = config.get("mode")
            if mode not in (1, 2):
                fatal.append(f"[{sheet_name}] 취합 방식이 올바르지 않습니다: {mode}")
            key_col = config.get("key_col", "")
            if key_col and key_col not in self.KEY_OPTIONS[1:]:
                fatal.append(f"[{sheet_name}] 기준 열이 올바르지 않습니다: {key_col}")
            if not isinstance(config.get("protect"), bool):
                fatal.append(f"[{sheet_name}] 빈칸만 채우기 값이 논리값이 아닙니다.")

        metadata = profile.get("metadata", {})
        template_name = metadata.get("template_file_name")
        current_name = os.path.basename(self.template_path)
        if template_name and template_name != current_name:
            minor.append(
                f"기준 취합양식은 '{template_name}'이고 현재 파일은 '{current_name}'입니다."
            )
        if metadata.get("app_version") not in (None, VERSION):
            minor.append(
                f"프로파일 생성 버전은 {metadata.get('app_version')}이고 "
                f"현재 버전은 {VERSION}입니다."
            )
        if metadata.get("sheet_count") not in (None, len(configs)):
            minor.append(
                f"메타데이터의 시트 수({metadata.get('sheet_count')})와 "
                f"실제 설정 수({len(configs)})가 다릅니다."
            )
        return fatal, minor

    def _apply_profile(self, profile):
        for sheet_name, row in self.sheet_vars.items():
            row["check"].set(False)
            row["s"].set("1")
            row["e"].set("")
            row["mode"].set(self.MODE_OPTIONS[0])
            row["key"].set(self.KEY_OPTIONS[0])
            row["protect"].set(True)
            self._selection_changed(sheet_name)

        for config in profile["sheet_configs"]:
            sheet_name = config["sheet_name"]
            row = self.sheet_vars[sheet_name]
            row["check"].set(True)
            row["s"].set(str(config["header_start"]))
            row["e"].set(str(config["header_end"]))
            row["mode"].set(self.MODE_OPTIONS[int(config["mode"]) - 1])
            row["key"].set(config.get("key_col") or self.KEY_OPTIONS[0])
            row["protect"].set(bool(config["protect"]))
            self._selection_changed(sheet_name)
        first_sheet = profile["sheet_configs"][0]["sheet_name"]
        self.select_sheet(first_sheet)

    def load_profile(self):
        initial_dir = profile_directory(SHEET_PROFILE_TYPE)
        path = filedialog.askopenfilename(
            parent=self.top,
            title="Sheet 설정 프로파일 불러오기",
            initialdir=initial_dir if os.path.isdir(initial_dir) else application_dir(),
            filetypes=[("Sheet 설정 프로파일", "*.json")],
        )
        if not path:
            return
        try:
            profile, _legacy = read_profile(path, SHEET_PROFILE_TYPE)
            fatal, minor = self._validate_profile(profile)
            if fatal:
                raise ValueError("\n".join(f"- {reason}" for reason in fatal))
            if minor and not messagebox.askyesno(
                PROGRAM_NAME,
                "프로파일과 현재 파일에 차이가 있습니다.\n\n"
                + "\n".join(f"- {reason}" for reason in minor)
                + "\n\n저장된 설정을 계속 적용하시겠습니까?",
                parent=self.top,
            ):
                return
            self._apply_profile(profile)
            profile_name = profile.get("metadata", {}).get("profile_name")
            self.profile_label_var.set(
                f"적용 프로파일: {profile_name or os.path.basename(path)}"
            )
            logging.info("Sheet 설정 프로파일 불러오기: %s, 경미한 불일치=%s", path, minor)
        except Exception as error:
            logging.error("Sheet 설정 프로파일 불러오기 실패: %s", error, exc_info=True)
            messagebox.showerror(
                PROGRAM_NAME,
                f"프로파일을 적용할 수 없습니다.\n\n사유:\n{error}",
                parent=self.top,
            )

    def save_profile(self, selected_sheets=None):
        selected_sheets = selected_sheets or self._collect_settings()
        if not selected_sheets:
            return False
        profile = self._profile_document(selected_sheets)
        default_dir = profile_directory(SHEET_PROFILE_TYPE)
        try:
            initial_dir = prepare_profile_directory(SHEET_PROFILE_TYPE)
        except OSError as error:
            logging.warning("Sheet 프로파일 폴더 준비 실패: %s", error)
            messagebox.showwarning(
                PROGRAM_NAME,
                "기본 프로파일 폴더에 저장할 수 없습니다.\n\n"
                f"경로: {default_dir}\n사유: {error}\n\n다른 저장 위치를 선택해 주세요.",
                parent=self.top,
            )
            initial_dir = os.path.dirname(self.template_path)
        path = filedialog.asksaveasfilename(
            parent=self.top,
            title="Sheet 설정 프로파일 저장",
            initialdir=initial_dir,
            defaultextension=".json",
            filetypes=[("Sheet 설정 프로파일", "*.json")],
        )
        if not path:
            return False
        profile["metadata"]["profile_name"] = os.path.splitext(
            os.path.basename(path)
        )[0]
        try:
            write_profile(profile, path)
            self.profile_label_var.set(
                f"적용 프로파일: {profile['metadata']['profile_name']}"
            )
            logging.info("Sheet 설정 프로파일 저장: %s, 시트=%s", path, len(selected_sheets))
            messagebox.showinfo(
                PROGRAM_NAME,
                f"Sheet 설정 프로파일을 저장했습니다.\n\n{path}",
                parent=self.top,
            )
            return True
        except Exception as error:
            logging.error("Sheet 설정 프로파일 저장 실패: %s", error, exc_info=True)
            messagebox.showerror(
                PROGRAM_NAME,
                f"프로파일을 저장하지 못했습니다.\n\n경로: {path}\n사유: {error}",
                parent=self.top,
            )
            return False

    def on_ok(self):
        selected_sheets = self._collect_settings()
        if not selected_sheets:
            return
        if messagebox.askyesno(
            PROGRAM_NAME,
            "현재 Excel 시트 설정을 프로파일로 저장하시겠습니까?",
            parent=self.top,
        ):
            self.save_profile(selected_sheets)
        self.selected_sheets = selected_sheets
        self._close()

    def _close(self):
        if self._closed:
            return
        self._closed = True
        if self._preview_after_id is not None:
            try:
                self.top.after_cancel(self._preview_after_id)
            except tk.TclError:
                pass
            self._preview_after_id = None
        if self.workbook is not None:
            self.workbook.close()
            self.workbook = None
        try:
            self.top.grab_release()
        except tk.TclError:
            pass
        self.top.destroy()

    def on_cancel(self):
        self.selected_sheets = {}
        self._close()

# GUI는 입력 선택과 결과 알림만 맡고, 실제 취합은 services.sheet_service가 담당한다.
def _open_result_folder(output_path):
    try:
        os.startfile(os.path.dirname(os.path.abspath(output_path)))
        return True
    except OSError as error:
        logging.warning("Excel 결과 폴더 열기 실패: %s", error)
        return False
# 이 정의는 위의 기존 GUI 중심 구현을 호환용으로 보존하면서 런타임 진입점을 교체한다.
def run_application(parent_root):
    from services.sheet_service import merge_workbooks

    logging.info("=== %s Excel Sheet 취합 GUI 시작 ===", PROGRAM_NAME)
    template_path = filedialog.askopenfilename(
        parent=parent_root,
        title="1. 데이터 취합 템플릿을 선택하세요",
        filetypes=[("Excel Files", "*.xlsx *.xlsm *.xls"), ("All Files", "*.*")],
    )
    if not template_path:
        return
    if template_path.lower().endswith(".xls"):
        messagebox.showerror(PROGRAM_NAME, "템플릿은 .xlsx 또는 .xlsm 파일이어야 합니다.", parent=parent_root)
        return
    try:
        workbook = load_workbook(template_path, read_only=True)
        sheet_names = workbook.sheetnames
        workbook.close()
        selector = MultiSheetSelector(parent_root, template_path, sheet_names)
        settings = selector.selected_sheets
        if not settings:
            return
        source_paths = filedialog.askopenfilenames(
            parent=parent_root,
            title="2. 취합할 원본 Excel 파일을 선택하세요",
            filetypes=[("Excel Files", "*.xlsx *.xlsm *.xls")],
        )
        if not source_paths:
            return
        from processing_cancellation import ProcessingCancellation

        cancellation = ProcessingCancellation()
        progress = ProcessingProgressDialog(
            parent_root,
            "Excel 취합 진행 상황",
            len(source_paths),
            "파일",
            cancellation=cancellation,
        )

        def report(event):
            progress.update_from_event(event)

        result = merge_workbooks(
            template_path,
            source_paths,
            None,
            os.path.dirname(template_path),
            report,
            settings=settings,
            cancellation=cancellation,
        )
        progress.close()
        if result.state.value == "cancelled":
            messagebox.showinfo(
                "취소됨",
                "전체 작업이 취소되었습니다.\n이번 작업의 결과 파일은 생성되지 않았습니다.",
                parent=parent_root,
            )
            return False
        if result.output_files:
            messagebox.showinfo(PROGRAM_NAME, f"Excel 취합이 완료되었습니다.\n\n{result.output_files[0]}", parent=parent_root)
            _open_result_folder(result.output_files[0])
        else:
            messagebox.showwarning(PROGRAM_NAME, result.message or "취합할 데이터가 없습니다.", parent=parent_root)
    except Exception as error:
        logging.exception("Excel Sheet 취합 실패")
        messagebox.showerror(PROGRAM_NAME, f"오류가 발생했습니다.\n\n{error}", parent=parent_root)
