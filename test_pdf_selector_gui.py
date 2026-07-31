"""PDF 설정 창의 비동기 로딩·동기화·미리보기 GUI 연기시험."""

import json
import tempfile
import time
import tkinter as tk
from pathlib import Path

from openpyxl import Workbook

from engine_Drag import SheetGroupSelector


def main():
    results = {}
    with tempfile.TemporaryDirectory(prefix="dns_auto_gui_test_") as temp_dir:
        workbook_path = Path(temp_dir) / "template.xlsx"
        workbook = Workbook()
        first = workbook.active
        first.title = "4월"
        second = workbook.create_sheet("5월")
        for worksheet in (first, second):
            worksheet.merge_cells("A2:A3")
            worksheet["A2"] = "기관"
            worksheet["B2"] = "처리"
            worksheet["B3"] = "결과"
            for row in range(4, 1004):
                worksheet.cell(row=row, column=1, value=row)
        workbook.save(workbook_path)
        workbook.close()

        root = tk.Tk()
        root.withdraw()
        original_finish = SheetGroupSelector._finish_workbook_load

        def finish_and_verify(selector, load_result):
            original_finish(selector, load_result)
            if load_result[0] != "ok":
                selector.on_cancel()
                return
            selector.rows["4월"]["start"].set("2")
            selector.rows["4월"]["end"].set("3")
            results["group_range_sync"] = (
                selector.rows["5월"]["start"].get() == "2"
                and selector.rows["5월"]["end"].get() == "3"
            )
            selector.select_sheet("4월")
            selector._refresh_preview()
            results["preview_row_count"] = len(selector.preview.get_children()) == 13
            results["workbook_loaded"] = selector.workbook is not None
            selector.root.after(50, selector.on_cancel)

        SheetGroupSelector._finish_workbook_load = finish_and_verify
        started = time.perf_counter()
        try:
            selector = SheetGroupSelector(str(workbook_path), parent_root=root)
            results["cancelled_cleanly"] = not selector.confirmed and selector.workbook is None
            results["elapsed_seconds"] = round(time.perf_counter() - started, 3)
        finally:
            SheetGroupSelector._finish_workbook_load = original_finish
            root.destroy()

    checks = {
        key: value
        for key, value in results.items()
        if key != "elapsed_seconds"
    }
    all_passed = all(checks.values())
    print(json.dumps({"results": results, "all_passed": all_passed}, ensure_ascii=False, indent=2))
    raise SystemExit(0 if all_passed else 1)


if __name__ == "__main__":
    main()
