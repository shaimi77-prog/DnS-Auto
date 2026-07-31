"""공용 시트 미리보기 범위·병합셀 로직의 비대화형 단위시험."""

import json

from openpyxl import Workbook

from utils_sheet_preview import (
    EXCEL_MAX_ROW,
    build_merged_value_lookup,
    preview_bounds,
)


def main():
    checks = {}
    checks["start_only_bounds"] = preview_bounds(20) == (10, 30)
    checks["start_end_bounds"] = preview_bounds(20, 25) == (10, 35)
    checks["lower_boundary"] = preview_bounds(5, 20) == (1, 30)
    checks["upper_boundary"] = preview_bounds(
        EXCEL_MAX_ROW - 2,
        EXCEL_MAX_ROW,
    ) == (EXCEL_MAX_ROW - 12, EXCEL_MAX_ROW)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.merge_cells("A2:C2")
    worksheet["A2"] = "총계"
    worksheet.merge_cells("A3:A14")
    worksheet["A3"] = "법무부"
    lookup = build_merged_value_lookup(
        worksheet,
        minimum_row=1,
        maximum_row=11,
        maximum_column=3,
    )
    checks["horizontal_merge"] = all(
        lookup[(2, column)] == "총계" for column in range(1, 4)
    )
    checks["vertical_merge_clamped"] = (
        lookup[(3, 1)] == "법무부"
        and lookup[(11, 1)] == "법무부"
        and (12, 1) not in lookup
    )
    checks["column_limit"] = all(column <= 3 for _row, column in lookup)
    workbook.close()

    failures = [name for name, passed in checks.items() if not passed]
    print(json.dumps(
        {"checks": checks, "all_passed": not failures},
        ensure_ascii=False,
        indent=2,
    ))
    raise SystemExit(1 if failures else 0)


if __name__ == "__main__":
    main()
